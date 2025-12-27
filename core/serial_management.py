# mogno_app/core/serial_management.py
import os
import pandas as pd
from openpyxl import load_workbook
from utils.logger import adicionar_log
import unicodedata

_serials_list = []
_arquivo_carregado = False
_origem_serials = None  # 'arquivo' ou 'manual'

# -------------------------
# Helpers CSV
# -------------------------
def _detectar_header_e_delimitador_csv(filepath, max_linhas=20):
    """
    Lê as primeiras max_linhas do CSV e tenta detectar:
    - possível linha de header (se conter palavras-chave)
    - delimitador mais provável (, ; \t)
    Retorna: (header_row_or_None (0-based), sep_or_None)
    """
    palavras_chave = [
        "serial", "seriais", "numero_serial", "número de série",
        "numero de serie", "serial_number", "nserie", "sn"
    ]
    palavras_chave = [p.lower() for p in palavras_chave]

    linhas = []
    try:
        with open(filepath, "r", errors="ignore") as f:
            for _ in range(max_linhas):
                try:
                    linhas.append(next(f).rstrip("\n"))
                except StopIteration:
                    break
    except Exception:
        return None, None

    candidatos = [",", ";", "\t"]
    sep_scores = {s: 0 for s in candidatos}
    for linha in linhas:
        for s in candidatos:
            sep_scores[s] += len(linha.split(s))
    detected_sep = max(candidatos, key=lambda s: sep_scores[s]) if any(v > 0 for v in sep_scores.values()) else None

    for idx, linha in enumerate(linhas):
        parts = [p.strip() for p in (linha.split(detected_sep) if detected_sep else linha.split(","))]
        if len(parts) <= 1:
            parts = [p.strip() for p in linha.split(";")]
        if len(parts) <= 1:
            parts = [p.strip() for p in linha.split("\t")]
        for texto in parts:
            if any(kw in str(texto).lower() for kw in palavras_chave):
                return idx, detected_sep

    return None, detected_sep


def _ler_csv_com_codificacao_automatica(filepath, nome_arquivo, header_row=None, sep=None):
    """
    Lê CSV tentando diversas codificações. Lê tudo como string (dtype=str).
    """
    codificacoes_teste = ["utf-8", "utf-8-sig", "cp1252", "latin1"]
    ultima_excecao = None

    read_kwargs = {"dtype": str}
    if header_row is not None:
        read_kwargs["header"] = header_row
    if sep is not None:
        read_kwargs["sep"] = sep

    for encoding in codificacoes_teste:
        try:
            df = pd.read_csv(filepath, encoding=encoding, **read_kwargs)
            adicionar_log(f"Arquivo '{nome_arquivo}' lido com encoding: {encoding} (header={header_row}, sep={sep})")
            return df
        except Exception as e:
            ultima_excecao = e

    raise Exception(
        f"Falha ao ler '{nome_arquivo}' com as codificações testadas "
        f"({', '.join(codificacoes_teste)}). Erro final: {ultima_excecao}"
    )


# -------------------------
# Helpers Excel
# -------------------------
def _normalizar_texto(texto):
    """
    Remove acentos e normaliza texto para comparação.
    """
    if not texto:
        return ""
    # Remove acentos (NFD = Normalization Form Decomposed)
    nfd = unicodedata.normalize('NFD', texto)
    sem_acentos = ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')
    return sem_acentos.lower().strip()


def _localizar_cabecalho_excel(filepath, palavras_chave, max_linhas=300):
    """
    Busca a primeira ocorrência das palavras_chave na aba ativa do arquivo Excel
    usando openpyxl (respeita merged cells). Retorna:
      (header_row_0based, header_col_0based, cell_ref, sheet_name)
    ou (None, None, None, sheet_name)
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    sheet_name = ws.title
    palavras_chave_norm = [_normalizar_texto(p) for p in palavras_chave]

    max_busca = min(ws.max_row, max_linhas)
    adicionar_log(f"🔎 Buscando header em '{sheet_name}' (até linha {max_busca})...")

    for r in ws.iter_rows(min_row=1, max_row=max_busca, values_only=False):
        for cel in r:
            if cel is None or cel.value is None:
                continue

            texto_original = str(cel.value).strip()
            texto_lower = texto_original.lower()
            texto_norm = _normalizar_texto(texto_original)

            # Verifica se alguma palavra-chave está presente
            for kw_orig, kw_norm in zip(palavras_chave, palavras_chave_norm):
                if kw_norm in texto_norm or kw_orig.lower() in texto_lower:
                    cell_ref = cel.coordinate  # ex: 'P7'
                    row_idx = cel.row         # 1-based
                    col_idx = cel.column      # 1-based
                    adicionar_log(
                        f"✅ Palavra-chave '{kw_orig}' encontrada em '{sheet_name}' "
                        f"célula {cell_ref} (linha {row_idx}, coluna {col_idx}) "
                        f"valor: '{texto_original}'"
                    )
                    wb.close()
                    return row_idx - 1, col_idx - 1, cell_ref, sheet_name

    adicionar_log(f"⚠️ Nenhuma palavra-chave encontrada na aba '{sheet_name}'.")
    wb.close()
    return None, None, None, sheet_name


def _montar_df_excel_por_header(filepath, header_row_0based):
    """
    Lê o Excel com pandas (header=None) e monta um DataFrame cujo header é a linha header_row_0based.
    Lê os dados abaixo do header como strings (dtype=str).
    Remove linhas completamente vazias.
    Retorna DataFrame pronto e lista de colunas normalizadas.
    """
    df_raw = pd.read_excel(filepath, header=None, dtype=str)
    header_row = int(header_row_0based)
    header = df_raw.iloc[header_row].astype(str).fillna("").tolist()

    adicionar_log(f"📋 Header detectado na linha {header_row + 1}: {header[:15]}...")  # Primeiras 15 colunas

    # Normaliza nomes e torna únicos se necessário
    normalized = []
    seen = {}
    for i, v in enumerate(header):
        label = str(v).strip()
        if label == "" or label.lower().startswith("unnamed") or label.lower() == "nan":
            label = f"col_{i}"
        if label in seen:
            seen[label] += 1
            label = f"{label}_{seen[label]}"
        else:
            seen[label] = 0
        normalized.append(label)

    # Cria DataFrame com dados abaixo do header
    df = df_raw.iloc[header_row + 1:].copy().reset_index(drop=True)
    df.columns = normalized

    # Remove linhas completamente vazias
    linhas_antes = len(df)
    df = df.dropna(how='all').reset_index(drop=True)
    linhas_depois = len(df)

    if linhas_antes != linhas_depois:
        adicionar_log(f"🧹 Removidas {linhas_antes - linhas_depois} linhas vazias")

    adicionar_log(f"📊 DataFrame montado: {len(df)} linhas × {len(df.columns)} colunas")

    return df, normalized


# -------------------------
# Função principal
# -------------------------
def ler_arquivo_serials(filepath):
    """
    Lê arquivo .xlsx ou .csv, detecta a coluna de seriais e retorna:
    {'unicos': [...], 'total_lidos': N, 'duplicados': D}
    """
    global _serials_list, _arquivo_carregado, _origem_serials

    if not filepath:
        adicionar_log("Nenhum arquivo selecionado.")
        return {"unicos": [], "total_lidos": 0, "duplicados": 0}

    nome_arquivo = os.path.basename(filepath)
    extensao = os.path.splitext(filepath)[1].lower()

    try:
        coluna_serial = None
        cell_ref = None
        sheet_name = None
        header_col = None

        # -------------------------
        # XLSX path
        # -------------------------
        if extensao == ".xlsx":
            adicionar_log(f"📂 Arquivo '{nome_arquivo}' lido como Excel (.xlsx) — detectando header...")

            # Palavras-chave expandidas para melhor detecção
            palavras_chave = [
                "serial", "seriais", "numero_serial", "número de série",
                "numero de serie", "serial_number", "nserie", "sn",
                "num_serie", "n_serie", "serie", "equipment", "número serial"
            ]

            header_row, header_col, cell_ref, sheet_name = _localizar_cabecalho_excel(
                filepath, palavras_chave, max_linhas=300
            )

            if header_row is not None:
                # Monta df com header exato baseado na linha encontrada
                df, normalized_cols = _montar_df_excel_por_header(filepath, header_row)
                adicionar_log(f"✅ Header reconstruído a partir da linha {header_row + 1} na aba '{sheet_name}'")

                # Se header_col detectado, escolhe a coluna por índice posicional (seguro)
                if header_col is not None and header_col < len(normalized_cols):
                    coluna_serial = normalized_cols[header_col]
                    adicionar_log(f"🎯 Coluna de serial escolhida por posição: index {header_col} → '{coluna_serial}'")
            else:
                # Fallback: lê com header padrão do pandas
                df = pd.read_excel(filepath, dtype=str)
                df = df.dropna(how='all').reset_index(drop=True)
                adicionar_log("⚠️ Header não detectado — lendo com header padrão do pandas (dtype=str).")

        # -------------------------
        # CSV path
        # -------------------------
        elif extensao == ".csv":
            # Detecta header_row e sep nas primeiras linhas
            header_row_csv, detected_sep = _detectar_header_e_delimitador_csv(filepath, max_linhas=20)
            df = _ler_csv_com_codificacao_automatica(
                filepath, nome_arquivo, header_row=header_row_csv, sep=detected_sep
            )
            df = df.astype(str)
            df = df.dropna(how='all').reset_index(drop=True)

        else:
            raise ValueError(f"Formato de arquivo não suportado: {extensao}")

        if df.empty:
            raise ValueError("Arquivo vazio ou sem dados válidos.")

        # -------------------------
        # Se ainda não temos coluna_serial, tenta por rótulo (labels)
        # -------------------------
        if coluna_serial is None:
            palavras_chave_cols = [
                "serial", "seriais", "numero_serial", "número de série",
                "numero de serie", "serial_number", "nserie", "sn",
                "num_serie", "n_serie", "serie", "equipment", "número serial"
            ]
            palavras_chave_cols_norm = [_normalizar_texto(p) for p in palavras_chave_cols]

            for col in df.columns:
                col_norm = _normalizar_texto(str(col))
                col_lower = str(col).lower()

                for kw_orig, kw_norm in zip(palavras_chave_cols, palavras_chave_cols_norm):
                    if kw_norm in col_norm or kw_orig.lower() in col_lower:
                        coluna_serial = col
                        adicionar_log(f"🔑 Coluna de serial encontrada por nome: '{coluna_serial}'")
                        break
                if coluna_serial:
                    break

        # -------------------------
        # Se ainda não encontrou, tenta inferir por amostra de valores
        # -------------------------
        if coluna_serial is None:
            adicionar_log("🔍 Tentando inferir coluna de serial por análise de conteúdo...")
            amostra = df.head(100)  # Aumentado para 100 linhas
            melhor = (0.0, None, None)  # score, idx, col_label

            for idx, col in enumerate(amostra.columns):
                serie = amostra[col].dropna().astype(str)
                serie = serie[serie.str.strip() != ""]  # Remove strings vazias
                serie = serie[serie.str.lower() != "nan"]  # Remove "nan"

                if len(serie) < 5:  # Precisa de pelo menos 5 valores
                    continue

                # Critérios de scoring aprimorados
                curto = (serie.str.len() <= 30).mean()
                sem_espaco = (~serie.str.contains(r"\s", na=False)).mean()
                alfanum = serie.str.replace(r"[-_\.]", "", regex=True).str.isalnum().mean()
                nao_vazio = (serie.str.len() > 0).mean()

                # Penaliza colunas com muitos valores repetidos
                unicidade = serie.nunique() / len(serie) if len(serie) > 0 else 0

                score = (curto * 0.3) + (sem_espaco * 0.25) + (alfanum * 0.2) + (nao_vazio * 0.15) + (unicidade * 0.1)

                if score > melhor[0]:
                    melhor = (score, idx, col)
                    adicionar_log(f"  → Coluna '{col}' (idx {idx}): score {score:.3f}")

            if melhor[1] is not None and melhor[0] > 0.35:
                coluna_serial = melhor[2]
                adicionar_log(f"✅ Coluna de serial inferida por amostra: '{coluna_serial}' (score: {melhor[0]:.3f})")

        # -------------------------
        # Fallback final: primeira coluna com dados
        # -------------------------
        if coluna_serial is None:
            for col in df.columns:
                if df[col].notna().sum() > 0:
                    coluna_serial = col
                    adicionar_log(f"⚠️ Usando primeira coluna com dados: '{coluna_serial}'")
                    break

            if coluna_serial is None:
                coluna_serial = df.columns[0]
                adicionar_log(f"⚠️ Usando primeira coluna: '{coluna_serial}'")

        # Log final da coluna escolhida
        adicionar_log(
            f"🔑 Coluna de serial final: '{coluna_serial}' "
            f"(célula detectada: {cell_ref if cell_ref else 'N/A'})"
        )

        # -------------------------
        # Extração e limpeza (strings)
        # -------------------------
        serie_raw = df[coluna_serial].astype(str).str.strip()

        # Remove valores vazios, "nan", "None", etc.
        serie_raw = [
            s for s in serie_raw.tolist() 
            if s and s.lower() not in ["nan", "none", "", "null", "<na>"]
        ]

        total_lidos = len(serie_raw)

        # Mantém ordem e remove duplicados
        serials_unicos = list(dict.fromkeys(serie_raw))
        duplicados = total_lidos - len(serials_unicos)

        _serials_list = serials_unicos
        _arquivo_carregado = True
        _origem_serials = "arquivo"

        # LOG ÚNICO - removida duplicação
        adicionar_log(
            f"✅ Arquivo '{nome_arquivo}' processado: "
            f"{total_lidos} lidos, {duplicados} duplicados removidos, "
            f"{len(serials_unicos)} únicos armazenados."
        )

        return {
            "unicos": serials_unicos,
            "total_lidos": total_lidos,
            "duplicados": duplicados
        }

    except Exception as e:
        import traceback
        adicionar_log(f"❌ Erro ao ler '{nome_arquivo}': {e}")
        adicionar_log(f"Stack trace: {traceback.format_exc()}")
        _serials_list = []
        _arquivo_carregado = False
        _origem_serials = None
        return {"unicos": [], "total_lidos": 0, "duplicados": 0}


# -------------------------
# Inserção manual
# -------------------------
def carregar_seriais_manualmente(texto):
    global _serials_list, _arquivo_carregado, _origem_serials

    if not texto:
        adicionar_log("⚠️ Nenhum serial manual inserido.")
        return {"unicos": [], "total_lidos": 0, "duplicados": 0}

    partes = [s.strip() for s in texto.replace("\n", ";").split(";") if s.strip()]
    total_lidos = len(partes)
    serials_unicos = list(dict.fromkeys(partes))
    duplicados = total_lidos - len(serials_unicos)

    _serials_list = serials_unicos
    _arquivo_carregado = False
    _origem_serials = "manual"

    adicionar_log(f"✍️ Inserção manual: {total_lidos} lidos, {duplicados} duplicados removidos.")
    return {"unicos": serials_unicos, "total_lidos": total_lidos, "duplicados": duplicados}


# -------------------------
# Utilitários
# -------------------------
def get_seriais():
    return _serials_list.copy()


def limpar_seriais():
    global _serials_list, _arquivo_carregado, _origem_serials
    _serials_list = []
    _arquivo_carregado = False
    _origem_serials = None
    adicionar_log("🧹 Lista de serials limpa.")


def get_info_serials():
    return {
        "quantidade": len(_serials_list),
        "origem": _origem_serials,
        "arquivo_carregado": _arquivo_carregado
    }
