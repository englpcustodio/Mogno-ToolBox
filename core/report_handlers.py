# mogno_app/core/report_handlers.py
"""
Gerenciamento central de geração de relatórios (separados e consolidados).
Inclui formatação automática, aba de resumo com hiperlinks clicáveis e cabeçalhos estilizados.
"""

import os
import traceback
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from utils.logger import adicionar_log
import importlib

# Importa módulos de relatório
from reports import (
    report_last_position_API,
    report_device_status_maxtrack_redis,
    report_last_position_redis,
    report_traffic_data_redis
)

# Evita cache antigo
for m in [
    report_last_position_API,
    report_device_status_maxtrack_redis,
    report_last_position_redis,
    report_traffic_data_redis
]:
    importlib.reload(m)


class ReportHandler:
    """Gerencia a geração de relatórios (separados e consolidados)."""

    REPORT_MAP = {
        "last_position_api": report_last_position_API,
        "last_position_redis": report_last_position_redis,
        "status_equipment": report_device_status_maxtrack_redis,
        "data_consumption": report_traffic_data_redis
    }

    REPORT_LABELS = {
        "last_position_api": "📡 Últimas Posições - API Mogno",
        "last_position_redis": "📍 Últimas Posições - Redis",
        "status_equipment": "⚙️ Status dos Equipamentos",
        "data_consumption": "📶 Consumo de Dados no Servidor"
    }

    def __init__(self, app_state, signal_manager, main_window):
        self.app_state = app_state
        self.signal_manager = signal_manager
        self.main_window = main_window

    # -------------------------------------------------------------------------
    # RELATÓRIO SEPARADO
    # -------------------------------------------------------------------------
    def generate_separate_reports(self, options: dict):
        try:
            adicionar_log("📁 Iniciando geração de relatórios separados...")

            serials = options.get("serials", [])
            enabled = options.get("enabled_queries", [])

            base_dir = os.path.join(os.getcwd(), "relatorios_gerados")
            os.makedirs(base_dir, exist_ok=True)

            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

            subdir_map = {
                "last_position_api": "ultimas_posicoes",
                "last_position_redis": "ultimas_posicoes",
                "status_equipment": "status_equipamentos",
                "data_consumption": "consumo_dados"
            }

            for query in enabled:
                adicionar_log(f"📁 Solicitado relatório separado: {query}")

                module = self.REPORT_MAP.get(query)
                if not module:
                    adicionar_log(f"⚠️ Tipo de relatório desconhecido: {query}")
                    continue

                adicionar_log(f"📂 Módulo carregado de: {getattr(module, '__file__', 'desconhecido')}")

                if not hasattr(module, "gerar_relatorio"):
                    adicionar_log(
                        f"⚠️ O módulo '{module.__name__}' não possui a função gerar_relatorio. "
                        f"Funções disponíveis: {', '.join([x for x in dir(module) if not x.startswith('_')])}"
                    )
                    continue

                resultados = self.app_state.get("dados_atuais", {}).get(query, [])
                if not resultados:
                    adicionar_log(f"⚠️ Nenhum dado disponível para {query}. Ignorando.")
                    continue

                output_dir = os.path.join(base_dir, subdir_map.get(query, "outros"))
                os.makedirs(output_dir, exist_ok=True)

                filename = f"report_{query}_{timestamp}.xlsx"
                output_path = os.path.join(output_dir, filename)

                try:
                    adicionar_log(f"📁 Gerando '{filename}' em {os.path.relpath(output_dir)} ...")
                    rpath = module.gerar_relatorio(serials, resultados, output_path)
                    adicionar_log(f"✅ Relatório gerado: '{os.path.relpath(rpath or output_path)}'")

                except Exception as e:
                    adicionar_log(f"❌ Erro ao gerar relatório '{filename}': {e}")
                    adicionar_log(traceback.format_exc())

            self.signal_manager.show_toast_success.emit("✅ Relatórios separados gerados com sucesso!")

        except Exception as e:
            adicionar_log(f"❌ Erro inesperado em generate_separate_reports: {e}")
            adicionar_log(traceback.format_exc())

    # -------------------------------------------------------------------------
    # RELATÓRIO CONSOLIDADO
    # -------------------------------------------------------------------------
    def generate_consolidated_report(self, options: dict):
        try:
            adicionar_log("📊 Iniciando geração do relatório consolidado...")

            serials = options.get("serials", [])
            enabled = options.get("enabled_queries", [])

            base_dir = os.path.join(os.getcwd(), "relatorios_consolidados")
            os.makedirs(base_dir, exist_ok=True)

            wb = Workbook()
            summary_ws = wb.active
            summary_ws.title = "Resumo"

            # Cabeçalho
            summary_ws.append(["Relatório", "Descrição", "Total de Registros", "Link"])
            _formatar_cabecalho(summary_ws)

            any_sheet = False

            for query in enabled:
                module = self.REPORT_MAP.get(query)

                if not module:
                    adicionar_log(f"⚠️ Tipo de relatório desconhecido: {query}")
                    continue

                resultados = self.app_state.get("dados_atuais", {}).get(query, [])
                if not resultados:
                    adicionar_log(f"⚠️ Nenhum dado para {query}. Ignorando.")
                    continue

                sheet_name = self.REPORT_LABELS.get(query, query).replace("📡","").replace("📍","").replace("⚙️","").replace("📶","").strip()

                ws = wb.create_sheet(sheet_name[:31])

                # Preenchimento com segurança
                _preencher_aba(ws, resultados)

                summary_ws.append([
                    self.REPORT_LABELS.get(query, query),
                    f"Relatório consolidado: {sheet_name}",
                    len(resultados),
                    f"=HYPERLINK(\"#{sheet_name}!A1\";\"Abrir Aba\")"
                ])

                any_sheet = True

            if not any_sheet:
                adicionar_log("⚠️ Nenhum dado disponível para consolidar.")
                return

            # Estiliza aba Resumo
            _ajustar_colunas(summary_ws)
            _aplicar_estilo_zebra(summary_ws)
            summary_ws.freeze_panes = "A3"

            summary_ws.insert_rows(1)
            summary_ws["A1"] = "📘 Relatório Consolidado - Mogno Toolbox"
            summary_ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
            summary_ws.merge_cells("A1:D1")
            summary_ws["A1"].alignment = Alignment(horizontal="center")

            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            output_path = os.path.join(base_dir, f"relatorio_consolidado_{timestamp}.xlsx")
            wb.save(output_path)

            adicionar_log(f"✅ Consolidado salvo em: {os.path.relpath(output_path)}")
            self.signal_manager.show_toast_success.emit("✅ Relatório consolidado gerado com sucesso!")

        except Exception as e:
            adicionar_log(f"❌ ERRO em generate_consolidated_report: {e}")
            adicionar_log(traceback.format_exc())


# =============================================================================
# UTILITÁRIOS DE FORMATAÇÃO
# =============================================================================

def _preencher_aba(ws, resultados):
    """
    Insere dados na aba de forma segura SEM sobrescrever estilos criados
    pelos módulos report_*.
    """
    try:
        # Detecta se já há cabeçalho formatado pelo gerador
        ja_tem_cabecalho = ws.max_row > 0

        if not ja_tem_cabecalho:
            # Geramos cabeçalhos somente se o módulo report_* NÃO gerou
            if isinstance(resultados, list) and resultados and isinstance(resultados[0], dict):
                ws.append(list(resultados[0].keys()))
            else:
                ws.append(["Dados"])

            _formatar_cabecalho(ws)

        # Insere linhas
        if isinstance(resultados, dict):
            for k, v in resultados.items():
                ws.append([k, v])

        elif isinstance(resultados, list):
            if isinstance(resultados[0], dict):
                for r in resultados:
                    ws.append([r.get(k, "") for k in resultados[0].keys()])
            else:
                for r in resultados:
                    ws.append([r])

        # Ajuste visual somente se a aba é simples
        if not ja_tem_cabecalho:
            _aplicar_alinhamento(ws)
            _ajustar_colunas(ws)
            _aplicar_estilo_zebra(ws)

        ws.freeze_panes = "A2"

    except Exception as e:
        adicionar_log(f"❌ Erro em _preencher_aba: {e}")
        adicionar_log(traceback.format_exc())


def _formatar_cabecalho(ws):
    fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF")
    )

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")


def _ajustar_colunas(ws):
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, min(max_len + 3, 60))


def _aplicar_alinhamento(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")


def _aplicar_estilo_zebra(ws):
    fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                cell.fill = fill_even
