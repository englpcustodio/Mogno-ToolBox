import os
import pandas as pd
from datetime import datetime
from google.protobuf.json_format import MessageToDict
from utils.helpers import auto_size_columns
from utils.logger import adicionar_log
from openpyxl import load_workbook # Importar load_workbook para abrir o arquivo Excel
import json # Importar json para converter dicionários em strings JSON, se preferir


DIR_STATUS_MAXTRACK = 'relatorios_gerados/status_maxtrack'

def flatten_dict(data, parent_key=''):
    """Achata um dicionário, inclusive listas de objetos, com prefixos no estilo esperado."""
    items = {}
    for key, value in data.items():
        new_key = f"{parent_key}_{key}" if parent_key else key
        if isinstance(value, dict):
            items.update(flatten_dict(value, new_key))
        elif isinstance(value, list):
            for i, item in enumerate(value):
                if isinstance(item, dict):
                    items.update(flatten_dict(item, f"{new_key}_{i}"))
                else:
                    items[f"{new_key}_{i}"] = item
        else:
            items[new_key] = value
    return items

def relatorio_status_excel(seriais, dados_status):
    if not os.path.exists(DIR_STATUS_MAXTRACK):
        os.makedirs(DIR_STATUS_MAXTRACK)

    data_hora_formatada = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nome_arquivo = f'relatorio_status_{data_hora_formatada}.xlsx'
    caminho_arquivo = os.path.join(DIR_STATUS_MAXTRACK, nome_arquivo)

    registros_planilha_inicial = [] # Para a aba 'status_device_MXT'
    registros_planilha_detalhado = [] # Para a aba 'status_device_MXT_detalhado'

    for serial in seriais:
        status_data = next((item for item in dados_status if item['Serial'] == serial), None)
        
        if status_data and status_data.get("Dados"):
            try:
                dados_proto = status_data["Dados"]
                parsed_data = MessageToDict(dados_proto, preserving_proto_field_name=True)
                
                # --- MODIFICAÇÃO AQUI para a aba 'status_device_MXT' ---
                # Apenas duas colunas: 'Seriais' e 'Dados de Status' (dicionário como string)
                registro_inicial = {
                    "Seriais": serial,
                    "Dados de Status": json.dumps(parsed_data, ensure_ascii=False) # Converte o dict para string JSON
                    # Você pode usar str(parsed_data) se preferir uma representação de string mais simples
                }
                registros_planilha_inicial.append(registro_inicial)

                # Para a aba 'status_device_MXT_detalhado' (dados achatados) - SEM MUDANÇAS
                dados_achatados = flatten_dict(parsed_data)
                dados_achatados["Número de Série"] = serial
                registros_planilha_detalhado.append(dados_achatados)

            except Exception as e:
                # Adicionar erro para a planilha inicial
                registros_planilha_inicial.append({
                    "Seriais": serial,
                    "Dados de Status": f"Erro ao converter dados de status: {str(e)}"
                })
                # Adicionar erro para a planilha detalhada
                registros_planilha_detalhado.append({
                    "Número de Série": serial,
                    "Erro": f"Erro ao converter dados de status: {str(e)}"
                })
        else:
            # Adicionar status de "Não possui informações" para ambas as planilhas
            registros_planilha_inicial.append({
                "Seriais": serial,
                "Dados de Status": "Não possui informações de Status"
            })
            registros_planilha_detalhado.append({
                "Número de Série": serial,
                "Status": "Não possui informações de Status"
            })

    # Criar o DataFrame para os dados iniciais
    df_initial = pd.DataFrame(registros_planilha_inicial)
    # Garantir que 'Seriais' seja a primeira coluna (já deve ser pela forma como foi criado)
    if 'Seriais' in df_initial.columns:
        cols_initial = ['Seriais', 'Dados de Status'] # Definindo a ordem exata das colunas
        df_initial = df_initial[cols_initial]

    # Criar o DataFrame para os dados detalhados (achatados)
    df_detalhado = pd.DataFrame(registros_planilha_detalhado)
    # Ordena colunas deixando 'Número de Série' sempre na frente para df_detalhado
    if 'Número de Série' in df_detalhado.columns:
        cols_detalhado = ['Número de Série'] + [col for col in df_detalhado.columns if col != 'Número de Série']
        df_detalhado = df_detalhado[cols_detalhado]

    try:
        # Salvar em Excel com duas abas
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            # Aba para dados iniciais
            sheet_name_initial = 'status_device_MXT'
            df_initial.to_excel(writer, sheet_name=sheet_name_initial, index=False)
            
            # Aba para dados detalhados (achatados)
            sheet_name_detalhado = 'status_device_MXT_detalhado'
            df_detalhado.to_excel(writer, sheet_name=sheet_name_detalhado, index=False)

        # Abrir o arquivo Excel para ajustar as colunas
        book = load_workbook(caminho_arquivo)
        
        # Ajustar colunas da aba de dados iniciais
        if sheet_name_initial in book.sheetnames:
            sheet_initial = book[sheet_name_initial]
            auto_size_columns(sheet_initial)
        
        # Ajustar colunas da aba de dados detalhados
        if sheet_name_detalhado in book.sheetnames:
            sheet_detalhado = book[sheet_name_detalhado]
            auto_size_columns(sheet_detalhado)
        
        book.save(caminho_arquivo) # Salva o arquivo com as colunas ajustadas
        adicionar_log(f"📁 Relatório gerado: {caminho_arquivo}")
        return caminho_arquivo


    except Exception as e:
        print(f"Erro ao ajustar e salvar o arquivo Excel: {e}")
        print(f"O arquivo foi salvo, mas as colunas podem não estar ajustadas: {nome_arquivo}")

