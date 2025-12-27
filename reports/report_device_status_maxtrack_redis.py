# reports/report_device_status_maxtrack_redis.py
import os
import json
import pandas as pd
from datetime import datetime
from google.protobuf.json_format import MessageToDict
from openpyxl import load_workbook
from reports.reports_utils import auto_size_columns
from utils.logger import adicionar_log
import traceback

# -------------------------------------------------------------------------
# ACHATADOR DE DICIONÁRIOS / LISTAS
# -------------------------------------------------------------------------
def flatten_dict(data, parent_key=''):
    """Achata dicts e listas aninhadas."""
    items = {}
    for key, value in data.items():
        new_key = f"{parent_key}_{key}" if parent_key else key

        if isinstance(value, dict):
            items.update(flatten_dict(value, new_key))

        elif isinstance(value, list):
            for i, item in enumerate(value):
                if isinstance(item, dict):
                    items.update(flatten_dict(item, f"{new_key}_{i}"))
                else:
                    items[f"{new_key}_{i}"] = item

        else:
            items[new_key] = value

    return items

# -------------------------------------------------------------------------
# API para o ReportHandler
# -------------------------------------------------------------------------
def gerar_relatorio(serials, resultados, output_path):
    """Função pública chamada pelo handler."""
    try:
        adicionar_log("📘 [status_maxtrack] Iniciando geração do relatório...")
        caminho = relatorio_status_excel(serials, resultados, output_path)
        adicionar_log(f"✅ Relatório de status salvo em: {caminho}")
        return caminho

    except Exception as e:
        adicionar_log(f"❌ Erro fatal em gerar_relatorio(): {e}")
        adicionar_log(traceback.format_exc())
        return None

# -------------------------------------------------------------------------
# Função principal
# -------------------------------------------------------------------------
def relatorio_status_excel(seriais, dados_status, output_path):

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    registros_status_mxt = []
    registros_detalhados = []
    seriais_sem_status = []

    adicionar_log("🔍 Iniciando processamento serial por serial...")

    # Cabeçalhos fixos para a aba status_device_MXT
    colunas_fixas = [
        "Número de Série", "imei", "FT_FIRMWARE_APP", "FT_FIRMWARE_MODEM",
        "FT_BOOTLOADER", "FT_PROFILE", "FT_GEO_LIBRARY", "FT_CAN_LIBRARY",
        "FT_ACTIONS2", "FT_USERS", "FT_MAXIO_CONFIG", "FT_LORA_ACTIONS",
        "primaryICCID", "macBT", "lastResetReason", "loraID"
    ]

    for serial in seriais:

        try:
            status_data = next(
                (item for item in dados_status if item.get("Serial") == serial),
                None
            )

            # ---------------------------------------------------------
            # Sem dados → vai para aba separada
            # ---------------------------------------------------------
            if not status_data or not status_data.get("Dados"):
                adicionar_log(f"[{serial}] ⚠️ Sem dados de status no Redis.")
                seriais_sem_status.append({"Número de Série": serial})

                registros_detalhados.append({
                    "Número de Série": serial,
                    "Status": "Não possui informações de Status"
                })
                continue

            dados_brutos = status_data["Dados"]

            # ---------------------------------------------------------
            # Se for protobuf → converter com MessageToDict
            # ---------------------------------------------------------
            if hasattr(dados_brutos, "DESCRIPTOR"):
                adicionar_log(f"[{serial}] 🧩 Objeto protobuf detectado. Convertendo...")
                parsed_data = MessageToDict(
                    dados_brutos,
                    preserving_proto_field_name=True
                )

            # ---------------------------------------------------------
            # Pode vir bytes contendo JSON
            # ---------------------------------------------------------
            elif isinstance(dados_brutos, (bytes, bytearray)):
                adicionar_log(f"[{serial}] 📦 Bytes detectados. Tentando decodificar...")
                try:
                    parsed_data = json.loads(dados_brutos.decode("utf-8"))
                except:
                    adicionar_log(f"[{serial}] ❌ Bytes não eram JSON. Armazenando bruto...")
                    parsed_data = {"raw_bytes": str(dados_brutos)}

            # ---------------------------------------------------------
            # Pode vir dict já convertido
            # ---------------------------------------------------------
            elif isinstance(dados_brutos, dict):
                adicionar_log(f"[{serial}] 🔧 Dados já são dict. Usando diretamente.")
                parsed_data = dados_brutos

            # ---------------------------------------------------------
            # Pode ser string contendo JSON
            # ---------------------------------------------------------
            elif isinstance(dados_brutos, str):
                try:
                    parsed_data = json.loads(dados_brutos)
                except:
                    adicionar_log(f"[{serial}] ⚠️ String não era JSON. Salvando como texto.")
                    parsed_data = {"raw_string": dados_brutos}

            else:
                adicionar_log(f"[{serial}] ❓ Tipo desconhecido: {type(dados_brutos)}")
                parsed_data = {"raw_unknown": str(dados_brutos)}

            # ---------------------------------------------------------
            # Construção da linha formatada para status_device_MXT
            # ---------------------------------------------------------
            linha_status = {c: "" for c in colunas_fixas}
            linha_status["Número de Série"] = serial

            # imei
            linha_status["imei"] = (
                parsed_data.get("identificationPack", {}).get("imei", "")
            )

            # Campos diretos
            for chave in ["primaryICCID", "macBT", "lastResetReason", "loraID"]:
                linha_status[chave] = parsed_data.get(chave, "")

            # Arquivos FT_*
            files = parsed_data.get("files", [])

            for file in files:
                file_type = file.get("fileType", "")
                if file_type in linha_status:

                    # Caso major/minor/patch → firmware 3.1.22
                    if all(k in file for k in ("major", "minor", "patch")):
                        linha_status[file_type] = f"{file['major']}.{file['minor']}.{file['patch']}"

                    # Caso tenha fileID
                    elif "fileID" in file:
                        linha_status[file_type] = file["fileID"]

            # Adiciona à aba status_device_MXT (somente se tem dados)
            registros_status_mxt.append(linha_status)

            # Registro detalhado (achatar)
            dados_achatados = flatten_dict(parsed_data)
            dados_achatados["Número de Série"] = serial
            registros_detalhados.append(dados_achatados)

            adicionar_log(f"[{serial}] ✅ {len(dados_achatados)} campos achatados.")

        except Exception as e:
            adicionar_log(f"⚠️ Erro processando {serial}: {e}")
            adicionar_log(traceback.format_exc())

            # Serial com erro vai para sem_status
            seriais_sem_status.append({"Número de Série": serial})

            registros_detalhados.append({
                "Número de Série": serial,
                "Erro": f"Erro ao converter: {str(e)}"
            })

    # ---------------------------------------------------------------------
    # DataFrames
    # ---------------------------------------------------------------------
    df_status_mxt = pd.DataFrame(registros_status_mxt)
    df_sem_status = pd.DataFrame(seriais_sem_status)
    df_detalhado = pd.DataFrame(registros_detalhados)

    # Filtra df_detalhado para remover entradas sem dados
    df_detalhado_limpo = df_detalhado[
        ~(df_detalhado["Status"].astype(str).str.contains("Não possui informações de Status", na=False))
        if "Status" in df_detalhado.columns
        else df_detalhado
    ].copy()

    # Garantir ordem da coluna principal
    if "Número de Série" in df_detalhado_limpo.columns:
        cols = ["Número de Série"] + [c for c in df_detalhado_limpo.columns if c != "Número de Série"]
        df_detalhado_limpo = df_detalhado_limpo[cols]

    # ---------------------------------------------------------------------
    # Escrever Excel nas abas certas
    # ---------------------------------------------------------------------
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # Aba 1 — status_device_MXT com colunas fixas formatadas (SOMENTE com dados)
        if not df_status_mxt.empty:
            df_status_mxt.to_excel(writer, sheet_name="status_device_MXT", index=False)

        # Aba 2 — seriais sem status
        if not df_sem_status.empty:
            df_sem_status.to_excel(writer, sheet_name="sem_status", index=False)

        # Aba 3 — detalhado com conteúdo válido
        if not df_detalhado_limpo.empty:
            df_detalhado_limpo.to_excel(writer, sheet_name="status_device_MXT_detalhado", index=False)

    # Ajuste das colunas
    book = load_workbook(output_path)
    for aba in ["status_device_MXT", "sem_status", "status_device_MXT_detalhado"]:
        if aba in book.sheetnames:
            auto_size_columns(book[aba])

    book.save(output_path)

    adicionar_log(f"📁 Relatório final salvo em {output_path}")
    return output_path
