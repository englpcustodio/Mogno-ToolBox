# reports/reports_utils.py
"""
Utilitários compartilhados para geração de relatórios Excel.
Contém funções de formatação, parsing, conversão de datas e manipulação de workbooks.
"""

import os
import re
import traceback
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from utils.helpers import epoch_to_datetime
from utils.logger import adicionar_log

# =============================================================================
# PARSING E CONVERSÃO DE DADOS
# =============================================================================


def parse_dados_redis(dados_str):
    """
    Parser robusto para converter texto hierárquico do Redis em dict achatado.
    Mantém compatibilidade total com a versão original.

    Args:
        dados_str: String com dados hierárquicos (formato chave: valor)

    Returns:
        dict: Dicionário achatado com todas as chaves
    """
    try:
        dados = {}
        prefixos = []

        if not dados_str:
            return {}

        # Remove tags HTML e normaliza quebras de linha
        dados_str = re.sub(r'<[^>]*>', '', dados_str)
        dados_str = dados_str.replace("\r", "")
        dados_str = re.sub(r"\n+", "\n", dados_str.strip())

        for linha in dados_str.splitlines():
            linha = linha.strip()
            if not linha:
                continue

            # Fecha bloco hierárquico
            if linha == "}":
                if prefixos:
                    prefixos.pop()
                continue

            # Abre bloco hierárquico
            if linha.endswith("{"):
                prefixos.append(linha[:-1].strip())
                continue

            # Processa pares chave:valor
            if ":" in linha:
                try:
                    chave, valor = linha.split(":", 1)
                    chave = chave.strip()
                    valor = valor.strip().strip('"')

                    # Conversões automáticas
                    if isinstance(valor, str) and valor.lower() in ("true", "false"):
                        valor = valor.lower() == "true"
                    elif re.match(r"^-?\d+(\.\d+)?$", valor):
                        num = float(valor) if "." in valor else int(valor)
                        chave_l = chave.lower()

                        # Detectar timestamps e converter
                        if any(x in chave_l for x in ["data_hora_evento", "data_hora_recebimento", "data_hora_armazenamento"]):
                            dt = epoch_to_datetime(num)
                            if dt:
                                valor = dt.strftime("%d/%m/%Y - %H:%M:%S")
                            else:
                                valor = num
                        else:
                            valor = num

                    # Monta chave final com prefixos
                    chave_final = "_".join(prefixos + [chave]) if prefixos else chave
                    dados[chave_final] = valor

                except Exception as e:
                    adicionar_log(f"⚠️ parse_dados_redis: erro na linha '{linha}': {e}")
                    continue

        return dados

    except Exception as e:
        adicionar_log(f"❌ Falha crítica em parse_dados_redis(): {e}")
        return {}


def parse_date_value(val):
    """
    Converte diversos formatos de data em objeto datetime.

    Suporta:
    - Objetos datetime
    - Timestamps epoch (int/float, segundos ou milissegundos)
    - Strings formatadas "DD/MM/YYYY - HH:MM:SS"
    - Strings ISO

    Args:
        val: Valor a ser convertido

    Returns:
        datetime ou None
    """
    if val is None:
        return None

    if isinstance(val, datetime):
        return val

    try:
        # Timestamps numéricos
        if isinstance(val, (int, float)):
            return epoch_to_datetime(val)

        # Strings
        if isinstance(val, str):
            # Formato padrão do sistema
            try:
                return datetime.strptime(val, "%d/%m/%Y - %H:%M:%S")
            except:
                pass

            # ISO format
            try:
                return datetime.fromisoformat(val)
            except:
                pass

    except Exception:
        pass

    return None


# =============================================================================
# FORMATAÇÃO DE PLANILHAS EXCEL
# =============================================================================


def formatar_cabecalho(ws, row_num=1):
    """
    Aplica formatação premium ao cabeçalho de uma planilha.

    Args:
        ws: Worksheet do openpyxl
        row_num: Número da linha do cabeçalho (padrão: 1)
    """
    try:
        bold_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
            top=Side(style="thin", color="FFFFFF"),
            bottom=Side(style="thin", color="FFFFFF")
        )

        for cell in ws[row_num]:
            if cell.value:
                cell.font = bold_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

    except Exception as e:
        adicionar_log(f"⚠️ Erro em formatar_cabecalho: {e}")


def formatar_cabecalho_customizado(ws, row_num, cor_hex="305496"):
    """Formata cabeçalho com cor customizada."""
    try:
        bold_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
            top=Side(style="thin", color="FFFFFF"),
            bottom=Side(style="thin", color="FFFFFF")
        )

        for cell in ws[row_num]:
            if cell.value:
                cell.font = bold_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    except Exception as e:
        adicionar_log(f"⚠️ Erro em formatar_cabecalho_customizado: {e}")


def mesclar_e_formatar_celula(ws, range_str, texto, cor_fundo="D9E1F2", negrito=True):
    """
    Mescla células e aplica formatação.

    Args:
        ws: Worksheet
        range_str: Range de células (ex: "A1:D1")
        texto: Texto a inserir
        cor_fundo: Cor de fundo em hex
        negrito: Se True, aplica negrito
    """
    try:
        ws.merge_cells(range_str)
        cell = ws[range_str.split(':')[0]]
        cell.value = texto
        cell.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
        cell.font = Font(bold=negrito, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    except Exception as e:
        adicionar_log(f"⚠️ Erro em mesclar_e_formatar_celula: {e}")


def aplicar_estilo_zebra(ws, min_row=2):
    """
    Aplica estilo zebrado (linhas alternadas) ao corpo da planilha.

    Args:
        ws: Worksheet do openpyxl
        min_row: Linha inicial (padrão: 2, após cabeçalho)
    """
    try:
        zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for i, row in enumerate(ws.iter_rows(min_row=min_row), start=min_row):
            for cell in row:
                # Zebra nas linhas pares
                if i % 2 == 0:
                    cell.fill = zebra_fill

                cell.border = thin_border

                # Alinhamento inteligente
                if isinstance(cell.value, str) and len(cell.value) > 80:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    except Exception as e:
        adicionar_log(f"⚠️ Erro em aplicar_estilo_zebra: {e}")


def ajustar_largura_colunas(ws, min_width=10, max_width=80):
    """
    Ajusta automaticamente a largura das colunas baseado no conteúdo.

    Args:
        ws: Worksheet do openpyxl
        min_width: Largura mínima (padrão: 10)
        max_width: Largura máxima (padrão: 80)
    """
    try:
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass

            adjusted_width = max(min_width, min(max_length + 3, max_width))
            ws.column_dimensions[column_letter].width = adjusted_width

    except Exception as e:
        adicionar_log(f"⚠️ Erro em ajustar_largura_colunas: {e}")


def ajustar_largura_colunas_otimizado(ws, min_width=10, max_width=50, sample_size=1000):
    """
    Ajusta largura de colunas baseado em AMOSTRA dos dados (mais rápido).

    Args:
        ws: Worksheet do openpyxl
        min_width: Largura mínima
        max_width: Largura máxima
        sample_size: Número de linhas a amostrar (padrão: 1000)
    """
    try:
        num_rows = ws.max_row
        num_cols = ws.max_column

        if num_rows <= 1:
            return

        # Para planilhas grandes, amostra apenas N linhas
        if num_rows > sample_size:
            # Amostra: cabeçalho + primeiras 500 + últimas 500
            sample_rows = [1] + list(range(2, min(502, num_rows))) + list(range(max(num_rows - 499, 502), num_rows + 1))
            #adicionar_log(f"📏 Ajustando largura de {num_cols} colunas (amostra de {len(sample_rows)} linhas)")
        else:
            sample_rows = range(1, num_rows + 1)

        # Calcula largura por coluna
        for col_idx in range(1, num_cols + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for row_idx in sample_rows:
                try:
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass

            adjusted_width = max(min_width, min(max_length + 2, max_width))
            ws.column_dimensions[column_letter].width = adjusted_width

    except Exception as e:
        adicionar_log(f"⚠️ Erro em ajustar_largura_colunas_otimizado: {e}")


def auto_size_columns(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    # Converte para string para calcular o comprimento
                    cell_value_str = str(cell.value)
                    if len(cell_value_str) > max_length:
                        max_length = len(cell_value_str)
            except:
                pass
        adjusted_width = (max_length + 5)  # Adiciona um Padding
        sheet.column_dimensions[column_letter].width = adjusted_width


# =============================================================================
# CRIAÇÃO DE ABAS ESPECIALIZADAS
# =============================================================================


def criar_aba_detalhada_ordenada(wb, sheet_name, data_map, origem='redis'):
    """
    Cria aba detalhada com dados ordenados por data (mais recente primeiro).

    Args:
        wb: Workbook do openpyxl
        sheet_name: Nome da aba
        data_map: dict {serial: dados_raw}
        origem: 'redis' ou 'api'

    Returns:
        Worksheet criado ou None
    """
    if not data_map:
        return None

    try:
        rows = []
        all_keys = []

        # Processa cada serial
        for serial, raw in data_map.items():
            if origem == 'redis':
                if isinstance(raw, dict):
                    parsed = raw.copy()
                else:
                    parsed = parse_dados_redis(str(raw))
            else:  # API
                parsed = raw if isinstance(raw, dict) else {"valor": str(raw)}

            parsed["Serial"] = serial

            # Coleta chaves em ordem de descoberta
            for k in parsed.keys():
                if k not in all_keys:
                    all_keys.append(k)

            rows.append(parsed)

        # Monta cabeçalhos (Serial primeiro)
        headers = ["Serial"] + [k for k in all_keys if k != "Serial"]

        # Função para extrair datetime de um registro
        def extract_datetime(item):
            for hint in ["data_hora_evento", "data_hora_recebimento", "data_hora_armazenamento"]:
                # Busca exata
                if hint in item:
                    dt = parse_date_value(item.get(hint))
                    if dt:
                        return dt

                # Busca parcial (contém o hint)
                for k in item.keys():
                    if hint in k.lower():
                        dt = parse_date_value(item.get(k))
                        if dt:
                            return dt

            return datetime.fromtimestamp(0)  # Muito antigo como fallback

        # Ordena por data (mais recente primeiro)
        rows_sorted = sorted(rows, key=extract_datetime, reverse=True)

        # Cria aba e escreve dados
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)

        for item in rows_sorted:
            row = []
            for h in headers:
                v = item.get(h, "")
                # Trunca strings muito longas (limite Excel)
                if isinstance(v, str) and len(v) > 32000:
                    v = v[:32000] + "… [TRUNCADO]"
                row.append(v)
            ws.append(row)

        # Aplica formatação
        formatar_planilha_completa(ws)

        return ws

    except Exception as e:
        adicionar_log(f"❌ Erro em criar_aba_detalhada_ordenada para '{sheet_name}': {e}")
        return None


def criar_abas_por_periodo(wb, base_name, data_map, origem='redis', selected_periods=None):
    """
    Cria abas separadas por período (Hoje, 1-7, 8-15, +16 dias).

    Args:
        selected_periods: lista de períodos a incluir (ex: ["Hoje", "1-7"])
                         Se None, inclui todos
    """
    if not data_map:
        adicionar_log(f"ℹ️ Sem dados para {base_name}, nenhuma aba criada")
        return {}

    # ✅ CRÍTICO: Se None, usa todos (mas isso NÃO deve acontecer se o usuário usou o diálogo)
    if selected_periods is None:
        selected_periods = ["Hoje", "1-7", "8-15", "+16"]
        adicionar_log(f"⚠️ selected_periods=None, usando todos: {selected_periods}")

    adicionar_log(f"📊 criar_abas_por_periodo: {base_name}, períodos={selected_periods}")

    try:
        entries = []
        all_keys = []

        for serial, raw in data_map.items():
            if origem == 'redis':
                if isinstance(raw, dict):
                    parsed = raw.copy()
                else:
                    parsed = parse_dados_redis(str(raw))
            else:
                parsed = raw if isinstance(raw, dict) else {"valor": str(raw)}

            parsed["Serial"] = serial

            # Extrai data
            dt = None
            for hint in ["data_hora_evento", "data_hora_recebimento", "data_hora_armazenamento"]:
                if hint in parsed:
                    dt = parse_date_value(parsed.get(hint))
                    if dt:
                        break
                for k in parsed.keys():
                    if hint in k.lower():
                        dt = parse_date_value(parsed.get(k))
                        if dt:
                            break
                if dt:
                    break

            for k in parsed.keys():
                if k not in all_keys:
                    all_keys.append(k)

            entries.append({"serial": serial, "dt": dt, "data": parsed})

        # Agrupa por período
        ranges = {
            "Hoje": [],
            "1-7": [],
            "8-15": [],
            "+16": []
        }

        now = datetime.now()
        for e in entries:
            dt = e["dt"]
            if not dt:
                ranges["+16"].append(e)
                continue

            dias = (now - dt).days

            if dias == 0:
                ranges["Hoje"].append(e)
            elif 1 <= dias <= 7:
                ranges["1-7"].append(e)
            elif 8 <= dias <= 15:
                ranges["8-15"].append(e)
            elif dias >= 16:
                ranges["+16"].append(e)

        headers = ["Serial"] + [k for k in all_keys if k != "Serial"]
        created_sheets = {}

        for periodo, items in ranges.items():
            if not items:
                adicionar_log(f"ℹ️ {base_name}_{periodo}: sem dados")
                continue

            # >>> FILTRO PELAS ESCOLHAS DO USUÁRIO <<<
            if periodo not in selected_periods:
                adicionar_log(f"ℹ️ {base_name}_{periodo}: período NÃO selecionado, pulando")
                continue

            sheet_name = f"{base_name}_{periodo}"
            adicionar_log(f"✅ Criando aba: {sheet_name} ({len(items)} registros)")
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)

            items_sorted = sorted(
                items,
                key=lambda x: x["dt"] if x["dt"] else datetime.fromtimestamp(0),
                reverse=True
            )

            for item in items_sorted:
                row = []
                for h in headers:
                    v = item["data"].get(h, "")
                    if isinstance(v, str) and len(v) > 32000:
                        v = v[:32000] + "… [TRUNCADO]"
                    row.append(v)
                ws.append(row)

            formatar_planilha_completa(ws)
            created_sheets[periodo] = ws

        adicionar_log(f"✅ {len(created_sheets)} abas criadas para {base_name}")
        return created_sheets

    except Exception as e:
        adicionar_log(f"❌ Erro em criar_abas_por_periodo '{base_name}': {e}")
        adicionar_log(traceback.format_exc())
        return {}


# =============================================================================
# CARREGAMENTO DE REGRAS DE HW
# =============================================================================


def carregar_regras_hw(path=None):
    """
    Lê arquivo de regras de hardware.

    Formato esperado:
        Modelo | required: gsm | optional: p2p, lorawan

    Args:
        path: Caminho do arquivo de regras (se None, usa padrão)

    Returns:
        dict: {modelo: {"required": [...], "optional": [...]}}
    """
    if path is None:
        path = os.path.join(os.getcwd(), "assets", "model_hw_rules.txt")

    regras = {}

    if not os.path.exists(path):
        adicionar_log(f"⚠️ Arquivo de regras HW não encontrado: {path}")
        return regras

    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read().strip()

            if not content or all(line.strip().startswith("#") for line in content.splitlines()):
                adicionar_log(f"⚠️ Arquivo de regras HW vazio ou só com comentários: {path}")
                return regras

            for linha in content.splitlines():
                linha = linha.strip()

                if not linha or linha.startswith("#"):
                    continue

                try:
                    modelo, rest = linha.split("|", 1)
                    modelo = modelo.strip()
                    required = []
                    optional = []

                    partes = rest.split("|")
                    for p in partes:
                        p = p.strip()
                        if p.startswith("required:"):
                            required = [
                                x.strip().lower()
                                for x in p.replace("required:", "").split(",")
                                if x.strip()
                            ]
                        elif p.startswith("optional:"):
                            optional = [
                                x.strip().lower()
                                for x in p.replace("optional:", "").split(",")
                                if x.strip()
                            ]

                    regras[modelo] = {"required": required, "optional": optional}

                except Exception as e:
                    adicionar_log(f"⚠️ Erro lendo linha de regras HW (ignorada): '{linha}' - {e}")

        adicionar_log(f"✅ Regras HW carregadas: {len(regras)} modelos de '{path}'")

    except Exception as e:
        adicionar_log(f"❌ Erro ao carregar regras HW de '{path}': {e}")

    return regras


# =============================================================================
# FORMATAÇÃO OTIMIZADA PARA GRANDES VOLUMES
# =============================================================================


def formatar_planilha_modo_rapido(ws, header_row=1, cor_header="305496"):
    """
    Formatação leve para grandes volumes: apenas cabeçalho + filtro + freeze.
    NÃO aplica zebra nem ajuste de largura célula-a-célula.

    Args:
        ws: Worksheet do openpyxl
        header_row: Linha do cabeçalho (padrão: 1)
        cor_header: Cor do cabeçalho em hex (padrão: azul)
    """
    try:
        if ws.max_row < 1:
            return

        # Formata apenas o cabeçalho
        formatar_cabecalho_customizado(ws, header_row, cor_header)

        # Congela painéis
        ws.freeze_panes = f"A{header_row + 1}"

        # Autofilter
        if ws.max_column >= 1 and ws.max_row >= header_row:
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"

        adicionar_log(f"✅ Formatação rápida aplicada à aba '{ws.title}' ({ws.max_row} linhas)")

    except Exception as e:
        adicionar_log(f"⚠️ Erro em formatar_planilha_modo_rapido: {e}")


# =============================================================================
# FORMATAÇÃO OTIMIZADA PARA CONSULTAS MENORES
# =============================================================================


def formatar_planilha_completa(ws, header_row=1, aplicar_zebra=True, ajustar_colunas=True):
    """
    Aplica formatação completa (cabeçalho + corpo + ajustes) a uma planilha.

    Args:
        ws: Worksheet do openpyxl
        header_row: Linha do cabeçalho (padrão: 1)
        aplicar_zebra: Se True, aplica estilo zebrado (padrão: True)
        ajustar_colunas: Se True, ajusta largura das colunas (padrão: True)
    """
    try:
        if ws.max_row < 1:
            return

        # Formata cabeçalho
        formatar_cabecalho(ws, header_row)

        # Formata corpo (opcional)
#        if aplicar_zebra:
#            aplicar_estilo_zebra(ws, min_row=header_row + 1)


        # Congela painéis
        ws.freeze_panes = f"A{header_row + 1}"

        # Autofilter (mantém filtro nas colunas)
        if ws.max_column >= 1 and ws.max_row >= header_row:
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"

        # Ajustes finais (ao final do processo)
        if ajustar_colunas:
            ajustar_largura_colunas(ws)

    except Exception as e:
        adicionar_log(f"⚠️ Erro em formatar_planilha_completa: {e}")


# =============================================================================
# FORMATAÇÃO PARA A PLANILHA DE TRÁFEGO DE DADOS NO SERVIDOR
# =============================================================================


def formatar_planilha_consumo(ws, incluir_zebra=True):
    """
    Aplica formatação padrão para abas de consumo de dados.

    Características:
    - Cabeçalho azul com fonte branca
    - Zebra (linhas alternadas cinzas)
    - Cores de alerta: Vermelho (>50MB) e Amarelo (<1MB)
    - Bordas em todas as células
    - Largura automática de colunas
    - Congelamento da linha de cabeçalho

    Args:
        ws: Worksheet do openpyxl
        incluir_zebra (bool): Se True, aplica padrão zebra
    """
    try:
        # Estilos
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")     # >50MB
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # <1MB

        # ─────────────────────────────────────────────────────────────────
        # FORMATAÇÃO DO CABEÇALHO (linha 1)
        # ─────────────────────────────────────────────────────────────────
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # ─────────────────────────────────────────────────────────────────
        # FORMATAÇÃO DAS LINHAS DE DADOS
        # ─────────────────────────────────────────────────────────────────
        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=ws.max_column), start=2):
            try:
                # Obtém valor de MB (coluna 4 = índice 3)
                mb_value = row[3].value if len(row) > 3 else None

                # Determina cor da linha
                if isinstance(mb_value, (float, int)):
                    if mb_value > 50:
                        row_fill = red_fill
                    elif mb_value < 1:
                        row_fill = yellow_fill
                    else:
                        row_fill = zebra_fill if (incluir_zebra and i % 2 == 0) else None
                else:
                    row_fill = zebra_fill if (incluir_zebra and i % 2 == 0) else None

                # Aplica formatação a cada célula
                for cell in row:
                    if row_fill:
                        cell.fill = row_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

            except Exception as e:
                adicionar_log(f"⚠️ Erro ao formatar linha {i}: {e}")

        # ─────────────────────────────────────────────────────────────────
        # AJUSTE DE LARGURA DE COLUNAS
        # ─────────────────────────────────────────────────────────────────
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 3, 60))

        # ─────────────────────────────────────────────────────────────────
        # CONGELAMENTO E FILTROS
        # ─────────────────────────────────────────────────────────────────
        ws.freeze_panes = "A2"

        adicionar_log(f"✅ Aba '{ws.title}' formatada com sucesso")

    except Exception as e:
        adicionar_log(f"❌ Erro ao formatar planilha '{ws.title}': {e}")