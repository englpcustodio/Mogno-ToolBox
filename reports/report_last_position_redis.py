# reports/report_last_position_redis.py
"""
Report generator: Últimas Posições (Redis)
Opção B: código completo com comentários explicativos.

Principais funcionalidades:
- Preserva parser original (parse_dados) que converte texto hierárquico em dicionário achatado.
- Lê regras de HW em assets/modelos_hw_rules.txt para validação "inteligente".
- Gera aba 'Resumo_Tipos' com:
    Total de Seriais | N
    Tipo | Encontrado | Esperado | Cobertura (%)
    (Tabela de modelos e contagem)
    Depois a tabela principal: Serial | Modelo de HW | Posição GSM | Data/Hora GSM | ...
- Cria aba 'Equip_sem_posicao' como segunda aba (sem coluna Modelo de HW).
- Cria abas detalhadas GSM_Detalhado, LoRaWAN_Detalhado, P2P_Detalhado (colunas separadas),
  ordenadas por data_hora_evento do mais recente ao mais antigo.
- Cria abas por faixa de dias (Hoje, 1-7, 8-15, +16) para cada protocolo.
- Aplica autofilter em todas as abas com cabeçalho.
- Logs em pontos-chaves para facilitar debug.
"""

import os
import re
import traceback
import openpyxl
from datetime import datetime, date
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from utils.helpers import epoch_to_datetime
from utils.logger import adicionar_log

# Path para regras de modelos (arquivo editável pelo usuário)
PATH_MODEL_RULES = os.path.join(os.getcwd(), "assets", "modelos_hw_rules.txt")

# ----------------------------
# Leitura de regras de HW
# ----------------------------
def carregar_regras_hw(path=PATH_MODEL_RULES):
    """
    Lê um arquivo de regras com formato simples:
      Modelo | required: gsm | optional: p2p, lorawan
    Retorna dict: { 'MXT-130': {'required': ['gsm'], 'optional': ['p2p','lorawan']}, ... }
    """
    regras = {}
    if not os.path.exists(path):
        adicionar_log(f"⚠ Arquivo de regras HW não encontrado: {path}")
        return regras

    try:
        with open(path, "r", encoding="utf-8") as f:
            for linha in f:
                linha = linha.strip()
                if not linha or linha.startswith("#"):
                    continue
                try:
                    modelo, rest = linha.split("|", 1)
                    modelo = modelo.strip()
                    required = []
                    optional = []
                    partes = rest.split("|")
                    for p in partes:
                        p = p.strip()
                        if p.startswith("required:"):
                            required = [x.strip().lower() for x in p.replace("required:", "").split(",") if x.strip()]
                        elif p.startswith("optional:"):
                            optional = [x.strip().lower() for x in p.replace("optional:", "").split(",") if x.strip()]
                    regras[modelo] = {"required": required, "optional": optional}
                except Exception:
                    adicionar_log(f"⚠ Erro lendo linha regras HW (ignorada): {linha}")
        adicionar_log(f"✅ Regras HW carregadas: {len(regras)} modelos")
    except Exception as e:
        adicionar_log(f"❌ Erro ao carregar regras HW: {e}")
    return regras

REGRAS_HW = carregar_regras_hw()

# ----------------------------
# Parser (preservado e robusto)
# ----------------------------
def parse_dados(dados_str):
    """
    Parser robusto para converter texto hierárquico em dict achatado.
    Mantido compatível com sua versão funcional.
    """
    try:
        dados = {}
        prefixos = []
        if not dados_str:
            return {}

        # remove tags HTML
        dados_str = re.sub(r'<[^>]*>', '', dados_str)
        dados_str = dados_str.replace("\r", "")
        dados_str = re.sub(r"\n+", "\n", dados_str.strip())

        for linha in dados_str.splitlines():
            linha = linha.strip()
            if not linha:
                continue

            if linha == "}":
                if prefixos:
                    prefixos.pop()
                continue

            if linha.endswith("{"):
                prefixos.append(linha[:-1].strip())
                continue

            if ":" in linha:
                try:
                    chave, valor = linha.split(":", 1)
                    chave = chave.strip()
                    valor = valor.strip().strip('"')

                    # conversões
                    if isinstance(valor, str) and valor.lower() in ("true", "false"):
                        valor = valor.lower() == "true"
                    elif re.match(r"^-?\d+(\.\d+)?$", valor):
                        num = float(valor) if "." in valor else int(valor)
                        chave_l = chave.lower()
                        # detectar timestamps
                        if ("data_hora_evento" in chave_l
                            or "data_hora_recebimento" in chave_l
                            or "data_hora_armazenamento" in chave_l):
                            dt = epoch_to_datetime(num)
                            if dt:
                                valor = dt.strftime("%d/%m/%Y - %H:%M:%S")
                            else:
                                valor = num
                        else:
                            valor = num

                    chave_final = "_".join(prefixos + [chave]) if prefixos else chave
                    dados[chave_final] = valor
                except Exception as e:
                    adicionar_log(f"⚠️ parse_dados: erro na linha '{linha}': {e}")
                    continue
        return dados
    except Exception as e:
        adicionar_log(f"❌ Falha crítica em parse_dados(): {e}")
        adicionar_log(traceback.format_exc())
        return {}

# ----------------------------
# Utilitários
# ----------------------------
def _formatar_planilha(ws):
    """Aplica cabeçalho, zebra, bordas e freeze; também ajusta colunas."""
    try:
        bold = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
        # Cabeçalho
        if ws.max_row < 1:
            return
        for c in ws[1]:
            c.font = bold
            c.fill = header_fill
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
        # Linhas
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for c in row:
                if i % 2 == 0:
                    c.fill = zebra
                c.border = border
                # centralização só quando razoável; grandes textos mantêm left align
                if isinstance(c.value, str) and len(c.value) > 80:
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")
        # Ajuste colunas
        for col in ws.columns:
            try:
                tamanho = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, min(tamanho + 3, 80))
            except Exception:
                pass
        # Freeze e autofilter (aplica se houver cabeçalho)
        try:
            ws.freeze_panes = "A2"
            if ws.max_column >= 1 and ws.max_row >= 1:
                last_col = get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
        except Exception:
            pass
    except Exception as e:
        adicionar_log(f"⚠️ Erro em _formatar_planilha: {e}")

def _parse_date_value(val):
    """
    Recebe valor possivelmente em:
    - datetime
    - int/float epoch (s or ms)
    - string formatted "DD/MM/YYYY - HH:MM:SS"
    Retorna datetime ou None
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        if isinstance(val, (int, float)):
            dt = epoch_to_datetime(val)
            return dt
        if isinstance(val, str):
            # Try pre-formatted
            try:
                return datetime.strptime(val, "%d/%m/%Y - %H:%M:%S")
            except Exception:
                # try common ISO-ish
                try:
                    return datetime.fromisoformat(val)
                except Exception:
                    return None
    except Exception:
        return None
    return None

# ----------------------------
# Cria aba detalhada com colunas separadas, e ordena por data_hora_evento (desc)
# ----------------------------
def _create_detailed_sheet_sorted(wb, sheet_name, serial_raw_data_map):
    """
    Gera uma aba detalhada com colunas separadas a partir do raw_data_map:
      serial_raw_data_map: dict { serial: raw_string_or_parsed_dict }
    - Se raw for string, usa parse_dados para extrair dict
    - Coleta todas keys em ordem de ocorrência para manter consistência
    - Ordena linhas pelo campo 'data_hora_evento' (se existir), do mais recente p/ antigo
    """
    if not serial_raw_data_map:
        return None
    rows = []
    all_keys = []
    for serial, raw in serial_raw_data_map.items():
        if isinstance(raw, dict):
            parsed = raw.copy()
        else:
            parsed = parse_dados(str(raw))
        parsed["Serial"] = serial
        # coleta chaves
        for k in parsed.keys():
            if k not in all_keys:
                all_keys.append(k)
        rows.append(parsed)

    # Determine header order: Serial first, depois as chaves em ordem de descoberta
    headers = ["Serial"] + [k for k in all_keys if k != "Serial"]

    # Sorting by data_hora_evento if present (field may be 'data_hora_evento' or with prefixes)
    def extract_dt(item):
        # Prefer explicit key 'data_hora_evento', fallback to any key containing that substring
        for key in ["data_hora_evento", "data_hora_recebimento", "data_hora_armazenamento"]:
            if key in item:
                dt = _parse_date_value(item.get(key))
                if dt:
                    return dt
        # try keys containing substring
        for k in item.keys():
            if "data_hora_evento" in k or "data_hora_recebimento" in k or "data_hora_armazenamento" in k:
                dt = _parse_date_value(item.get(k))
                if dt:
                    return dt
        return datetime.fromtimestamp(0)  # muito antigo

    rows_sorted = sorted(rows, key=extract_dt, reverse=True)

    # Cria sheet e escreve
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for item in rows_sorted:
        row = []
        for h in headers:
            v = item.get(h, "")
            # truncate long strings for excel safety
            if isinstance(v, str) and len(v) > 32000:
                v = v[:32000] + "… [TRUNCADO]"
            row.append(v)
        ws.append(row)

    return ws

# ----------------------------
# Função principal exposta: gerar_relatorio()
# ----------------------------
def gerar_relatorio(serials_list, resultados, output_path, separado=True):
    """
    Interface esperada pelo ReportHandler:
      serials_list: lista de seriais solicitados (ordem original)
      resultados: lista de registros (cada registro pode ser dict com keys 'Serial','Tipo','Dados', 'Modelo de HW', 'DataHora Evento', etc.)
      output_path: caminho completo solicitado pelo handler para salvar o arquivo
    Retorna caminho salvo (output_path) ou None em caso de erro.
    """
    try:
        adicionar_log("📍 [report_last_position_redis] iniciando geração...")
        if not resultados:
            adicionar_log("⚠️ Nenhum resultado disponível para gerar relatório.")
            return None

        # workbook
        wb = openpyxl.Workbook()
        ws_main = wb.active
        ws_main.title = "Resumo_Tipos"

        # estruturas intermediárias
        contagem = {"gsm": 0, "lorawan": 0, "p2p": 0}
        serial_data = {}     # info por serial: modelo, datas e flags
        dados_gsm = {}       # raw string or parsed dict by serial
        dados_lorawan = {}
        dados_p2p = {}
        modelo_counts = {}

        adicionar_log(f"ℹ️ Processando {len(resultados)} registros...")

        # Processar registros (mantendo parser/interpretacao original)
        for idx, resultado in enumerate(resultados):
            try:
                if not isinstance(resultado, dict):
                    adicionar_log(f"⚠️ Item #{idx} ignorado: não é dict")
                    continue

                # As chaves podem variar: aceitar variantes
                serial = resultado.get("Serial") or resultado.get("serial") or resultado.get("rastreador_numero_serie")
                tipo_raw = resultado.get("Tipo") or resultado.get("tipo") or resultado.get("tipo_evento") or ""
                tipo = str(tipo_raw).lower().strip()

                modelo_hw = resultado.get("Modelo de HW") or resultado.get("modelo_hw") or resultado.get("rastreador_versao_hardware") or "N/A"
                # normaliza modelo para uso em regras (mantém original para exibição)
                modelo_display = modelo_hw

                dados_raw = resultado.get("Dados", "") or resultado.get("raw_package") or resultado.get("raw") or ""
                # Data/hora evento: pode vir em 'DataHora Evento' ou 'data_hora_evento' etc.
                data_evt = resultado.get("DataHora Evento") or resultado.get("data_hora_evento") or resultado.get("data_hora_recebimento") or resultado.get("data_hora_armazenamento") or None

                if serial is None:
                    continue

                # if not present in serial_data, init
                if serial not in serial_data:
                    serial_data[serial] = {
                        "Modelo de HW": modelo_display,
                        "Data GSM": None,
                        "Data LoRaWAN": None,
                        "Data P2P": None,
                        "Tem GSM": False,
                        "Tem LOR": False,
                        "Tem P2P": False
                    }
                    modelo_counts[modelo_display] = modelo_counts.get(modelo_display, 0) + 1

                # normalize event date
                if isinstance(data_evt, (int, float)):
                    dt_obj = epoch_to_datetime(data_evt)
                    if dt_obj:
                        data_evt_fmt = dt_obj.strftime("%d/%m/%Y - %H:%M:%S")
                    else:
                        data_evt_fmt = str(data_evt)
                elif isinstance(data_evt, str):
                    # if parser already formatted the date, keep it; otherwise try to parse numeric inside
                    data_evt_fmt = data_evt
                elif isinstance(data_evt, datetime):
                    data_evt_fmt = data_evt.strftime("%d/%m/%Y - %H:%M:%S")
                else:
                    data_evt_fmt = None

                # classify by tipo (gsm / lora / p2p)
                if "gsm" in tipo:
                    serial_data[serial]["Data GSM"] = data_evt_fmt or serial_data[serial]["Data GSM"]
                    serial_data[serial]["Tem GSM"] = True
                    # Prefer guardar dados brutos; parse later para detalhado
                    dados_gsm[serial] = dados_raw
                    contagem["gsm"] += 1
                elif "lora" in tipo or "lorawan" in tipo or "lora" in tipo:
                    serial_data[serial]["Data LoRaWAN"] = data_evt_fmt or serial_data[serial]["Data LoRaWAN"]
                    serial_data[serial]["Tem LOR"] = True
                    dados_lorawan[serial] = dados_raw
                    contagem["lorawan"] += 1
                elif "p2p" in tipo:
                    serial_data[serial]["Data P2P"] = data_evt_fmt or serial_data[serial]["Data P2P"]
                    serial_data[serial]["Tem P2P"] = True
                    dados_p2p[serial] = dados_raw
                    contagem["p2p"] += 1
                else:
                    # tipo não identificado: armazenar no campo P2P como fallback (não contamos para contagem)
                    # optar por não contar para evitar inflar métricas
                    serial_data[serial]["Data P2P"] = serial_data[serial].get("Data P2P") or data_evt_fmt
                    dados_p2p.setdefault(serial, dados_raw)

            except Exception as e:
                adicionar_log(f"⚠️ Erro processando registro #{idx}: {e}")
                adicionar_log(traceback.format_exc())
                continue

        total_seriais = len(serials_list)
        adicionar_log(f"ℹ️ Total de seriais solicitados: {total_seriais}. Seriais com qualquer informação: {len(serial_data)}")

        # -----------------------------------------
        # Resumo Inteligente por Modelo (Regra A)
        # -----------------------------------------
        # Para cada modelo, calcula quantos 'expected' (esperados) de cada tipo
        # Ex.: se modelo MXT-130 required: gsm, expected_gsm += count(modelo)
        expected = {"gsm": 0, "lorawan": 0, "p2p": 0}
        # Calcular expected por modelo usando REGRAS_HW
        for modelo, qtd in modelo_counts.items():
            regras = REGRAS_HW.get(modelo)
            if regras:
                for req in regras.get("required", []):
                    if req in expected:
                        expected[req] += qtd
                # optional não contribui para expected
            else:
                # sem regra: assume que todos modelos podem ter todos — não incrementa expected (conservador)
                pass

        # If expected remains zero for a protocol, set expected to number of serials (fallback)
        for proto in ("gsm", "lorawan", "p2p"):
            if expected[proto] == 0:
                # fallback conservador: esperamos que todos possam ter (but this may be adjusted)
                expected[proto] = sum(modelo_counts.values())

        # -----------------------------------------
        # Montar aba Resumo_Tipos
        # -----------------------------------------
        def pct_str(found, expect):
            try:
                return f"{found} | {expect} | { (found/expect*100):.1f}%"
            except Exception:
                return f"{found} | {expect} | 0.0%"

        # Top lines: Total de Seriais
        ws_main.append(["Total de Seriais", total_seriais])
        # blank line for readability
        ws_main.append([])

        # Header for Tipo table
        ws_main.append(["Tipo", "Encontrado", "Esperado", "Cobertura (%)"])
        # GSM
        try:
            cov_gsm = (contagem["gsm"] / expected["gsm"] * 100) if expected["gsm"] else 0
        except Exception:
            cov_gsm = 0
        ws_main.append(["GSM", contagem["gsm"], expected["gsm"], f"{cov_gsm:.1f}%"])
        try:
            cov_lor = (contagem["lorawan"] / expected["lorawan"] * 100) if expected["lorawan"] else 0
        except Exception:
            cov_lor = 0
        ws_main.append(["LoRaWAN", contagem["lorawan"], expected["lorawan"], f"{cov_lor:.1f}%"])
        try:
            cov_p2p = (contagem["p2p"] / expected["p2p"] * 100) if expected["p2p"] else 0
        except Exception:
            cov_p2p = 0
        ws_main.append(["P2P", contagem["p2p"], expected["p2p"], f"{cov_p2p:.1f}%"])

        # Sem comunicação (count serials which have none of the flags)
        sem_comunic = len([s for s in serials_list if not serial_data.get(s, {}).get("Tem GSM") and not serial_data.get(s, {}).get("Tem LOR") and not serial_data.get(s, {}).get("Tem P2P")])
        ws_main.append(["Sem comunicação", sem_comunic, "-", f"{(sem_comunic/total_seriais*100) if total_seriais else 0:.1f}%"])

        ws_main.append([])

        # Modelo de HW counts (sorted alphabetically)
        ws_main.append(["Modelo de HW", "Quantidade"])
        for modelo in sorted(modelo_counts.keys(), key=lambda s: (str(s).lower())):
            ws_main.append([modelo, modelo_counts[modelo]])
        ws_main.append([])

        # Cabeçalho da tabela principal pedida
        headers = ["Serial", "Modelo de HW", "Posição GSM", "Data/Hora GSM", "Posição LoRaWAN", "Data/Hora LoRaWAN", "Posição P2P", "Data/Hora P2P"]
        ws_main.append(headers)

        # Preencher linhas: ordenadas por Modelo de HW (alfabética)
        # Construir lista de tuples (modelo, serial) e ordenar
        serials_sorted = sorted(serials_list, key=lambda s: (str(serial_data.get(s, {}).get("Modelo de HW", "")).lower(), str(s)))
        for s in serials_sorted:
            info = serial_data.get(s, {"Modelo de HW": "N/A", "Data GSM": None, "Data LoRaWAN": None, "Data P2P": None})
            pos_gsm = "SIM" if info.get("Tem GSM") else "NÃO"
            pos_lor = "SIM" if info.get("Tem LOR") else "NÃO"
            pos_p2p = "SIM" if info.get("Tem P2P") else "NÃO"
            row = [
                s,
                info.get("Modelo de HW", "N/A"),
                pos_gsm,
                info.get("Data GSM") or "N/A",
                pos_lor,
                info.get("Data LoRaWAN") or "N/A",
                pos_p2p,
                info.get("Data P2P") or "N/A"
            ]
            ws_main.append(row)

        # Formatar a aba principal
        _formatar_planilha(ws_main)

        # -----------------------------------------
        # Aba Equip_sem_posicao (segunda aba) - sem coluna Modelo de HW
        # -----------------------------------------
        try:
            ws_sem = wb.create_sheet("Equip_sem_posicao")
            ws_sem.append(["Serial"])  # coluna única como solicitado
            for s in serials_list:
                info = serial_data.get(s, {})
                if not info.get("Tem GSM") and not info.get("Tem LOR") and not info.get("Tem P2P"):
                    ws_sem.append([s])
            _formatar_planilha(ws_sem)
            # move this sheet to be the second sheet: index 1 (0-based is main sheet)
            wb._sheets.remove(ws_sem)
            wb._sheets.insert(1, ws_sem)
        except Exception as e:
            adicionar_log(f"⚠️ Erro criando Equip_sem_posicao: {e}")

        # -----------------------------------------
        # Abas detalhadas (ordenadas por data_hora_evento)
        # As abas devem conter colunas separadas (parser)
        # -----------------------------------------
        try:
            # GSM
            ws = _create_detailed_sheet_sorted(wb, "GSM_Detalhado", dados_gsm)
            if ws:
                _formatar_planilha(ws)
            # LoRaWAN
            ws = _create_detailed_sheet_sorted(wb, "LoRaWAN_Detalhado", dados_lorawan)
            if ws:
                _formatar_planilha(ws)
            # P2P
            ws = _create_detailed_sheet_sorted(wb, "P2P_Detalhado", dados_p2p)
            if ws:
                _formatar_planilha(ws)
        except Exception as e:
            adicionar_log(f"⚠️ Erro criando abas detalhadas: {e}")
            adicionar_log(traceback.format_exc())

        # -----------------------------------------
        # Abas por faixa de dias (usam as mesmas linhas das abas detalhadas,
        # então vamos reutilizar os parsed dicts e ordenar por data)
        # -----------------------------------------
        try:
            # Helper que extrai from detailed sheet data: but we have dados_{proto} as raw strings
            # Para simplicidade, vamos gerar as abas a partir dos dados_gsm/dados_lorawan/dados_p2p
            def criar_abas_range(proto_name, dados_map):
                # Build list of parsed entries with serial + data_hora_evento field (parsed)
                entries = []
                for serial, raw in dados_map.items():
                    parsed = parse_dados(str(raw))
                    # tentativa de achar uma chave de data
                    dt = None
                    for key in parsed.keys():
                        if "data_hora_evento" in key or "data_hora_recebimento" in key or "data_hora_armazenamento" in key:
                            dt = _parse_date_value(parsed.get(key))
                            parsed["_data_key"] = key
                            break
                    # fallback: use serial_data timestamps
                    if dt is None:
                        # check serial_data collected earlier
                        info = serial_data.get(serial, {})
                        dt_str = info.get(f"Data {proto_name}") if proto_name in ("GSM","LoRaWAN","P2P") else None
                        dt = _parse_date_value(dt_str)
                    entries.append({"serial": serial, "parsed": parsed, "dt": dt})

                # Now create ranges: Hoje (dias==0), 1-7, 8-15, +16
                ranges = {
                    "Hoje": [],
                    "1-7": [],
                    "8-15": [],
                    "+16": []
                }
                now = datetime.now()
                for e in entries:
                    dt = e["dt"]
                    if not dt:
                        continue
                    dias = (now - dt).days
                    if dias == 0:
                        ranges["Hoje"].append(e)
                    elif 1 <= dias <= 7:
                        ranges["1-7"].append(e)
                    elif 8 <= dias <= 15:
                        ranges["8-15"].append(e)
                    elif dias >= 16:
                        ranges["+16"].append(e)

                # for each range create sheet with same columns as detailed sheet (we'll flatten rows)
                for key, list_entries in ranges.items():
                    name = f"{proto_name}_{key}"
                    wsr = wb.create_sheet(name)
                    # if no entries, still create header
                    # Collect all keys across parsed items to make columns consistent
                    all_keys = ["Serial"]
                    for e in list_entries:
                        for k in e["parsed"].keys():
                            if k not in all_keys:
                                all_keys.append(k)
                    wsr.append(all_keys)
                    # sort entries by dt desc
                    list_entries_sorted = sorted(list_entries, key=lambda x: x["dt"] or datetime.fromtimestamp(0), reverse=True)
                    for e in list_entries_sorted:
                        row = []
                        parsed = e["parsed"]
                        for keyh in all_keys:
                            if keyh == "Serial":
                                row.append(e["serial"])
                            else:
                                v = parsed.get(keyh, "")
                                if isinstance(v, str) and len(v) > 32000:
                                    v = v[:32000] + "… [TRUNCADO]"
                                row.append(v)
                        wsr.append(row)
                    _formatar_planilha(wsr)

            criar_abas_range("GSM", dados_gsm)
            criar_abas_range("LoRaWAN", dados_lorawan)
            criar_abas_range("P2P", dados_p2p)
        except Exception as e:
            adicionar_log(f"⚠️ Erro criando abas por faixa de dias: {e}")
            adicionar_log(traceback.format_exc())

        # -----------------------------------------
        # Reorder sheets: ensure desired order:
        # 0: Resumo_Tipos (already)
        # 1: Equip_sem_posicao
        # 2.. : GSM_Detalhado, LoRaWAN_Detalhado, P2P_Detalhado, ranges...
        # We'll attempt to set main order where possible.
        try:
            # Build desired order list
            desired_prefix = ["Resumo_Tipos", "Equip_sem_posicao",
                              "GSM_Detalhado", "LoRaWAN_Detalhado", "P2P_Detalhado"]
            # gather others (like ranges and consolidated)
            others = [s.title for s in wb._sheets if s.title not in desired_prefix]
            new_order = []
            for name in desired_prefix:
                if name in [s.title for s in wb._sheets]:
                    new_order.append(name)
            new_order.extend(sorted(others))
            # Reorder sheets
            sheets_map = {s.title: s for s in wb._sheets}
            wb._sheets = [sheets_map[name] for name in new_order if name in sheets_map]
        except Exception:
            # Non-fatal
            pass

        # -----------------------------------------
        # Salvar arquivo
        # -----------------------------------------
        try:
            # Garantir diretório
            out_dir = os.path.dirname(output_path)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)
            wb.save(output_path)
            adicionar_log(f"✅ Relatório salvo em: {output_path}")
            return output_path
        except Exception as e:
            adicionar_log(f"❌ Falha ao salvar arquivo: {e}")
            adicionar_log(traceback.format_exc())
            return None

    except Exception as e:
        adicionar_log(f"❌ Erro grave em gerar_relatorio(): {e}")
        adicionar_log(traceback.format_exc())
        return None
