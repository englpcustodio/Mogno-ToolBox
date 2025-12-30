# reports/report_last_position.py
"""
Gerador Unificado de Relatório: Últimas Posições
Suporta dados de API Mogno e Redis com tratamento específico para cada origem.
Mantém formatações e layouts específicos para cada fonte.

Atualizado para:
- Total de Seriais Válidos
- Encontrada/Esperada [%] = Encontrado / Total Esperado * 100
- Posição Hoje, Posição 1-7 dias, Posição 8-15 dias, Posição +16 dias
- Separação de Data e Hora (GSM, LoRaWAN, P2P)
- Ordenação por prioridade: GSM (mais recente) -> P2P -> LoRaWAN
- Contagens por período baseadas nas abas por período geradas
- Ordenação da aba Equip_sem_posicao do menor para o maior
- NOVO: Aba "Resumo_Ultima_posicao" limpa (sem bloco de Modelo de HW)
- NOVO: Aba "Resumo_Modelo_HW" obrigatória (segunda posição)
"""

import os
import traceback
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

from core.app_state import AppState
from utils.logger import adicionar_log
from reports.reports_utils import (
    parse_dados_redis,
    carregar_regras_hw,
    parse_date_value,
    formatar_cabecalho,
    aplicar_estilo_zebra,
    ajustar_largura_colunas,
    formatar_planilha_completa,
    criar_aba_detalhada_ordenada,
    criar_abas_por_periodo
)


# =============================================================================
# GERAÇÃO ESPECÍFICA PARA REDIS
# =============================================================================

def gerar_relatorio_redis(serials_list, resultados, output_path, selected_periods=None):
    """
    Gera relatório completo para dados do Redis.
    Mantém toda a lógica original com abas de resumo, detalhadas e por período.

    Args:
        serials_list: Lista de seriais solicitados
        resultados: Lista de registros com dados
        output_path: Caminho completo para salvar o arquivo
        selected_periods: Lista de períodos a incluir (ex: ["Hoje", "1-7"])
    """
    try:
        adicionar_log("📍 [Redis] Iniciando geração do relatório...")

        # Carrega regras HW (agora usando função centralizada)
        REGRAS_HW = carregar_regras_hw()

        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "Resumo_Ultima_posicao"

        # Estruturas de dados
        # contagem por serial (único): quantos seriais trouxeram posição por tecnologia
        found_serials = {"gsm": set(), "lorawan": set(), "p2p": set()}
        serial_data = {}
        dados_gsm = {}
        dados_lorawan = {}
        dados_p2p = {}
        modelo_counts = {}

        # Inicializa todos os seriais
        for s in serials_list:
            serial_data[s] = {
                "Modelo de HW": "N/A",
                "dt_gsm": None,
                "dt_lora": None,
                "dt_p2p": None,
                "Data GSM": None,
                "Hora GSM": None,
                "Data LoRaWAN": None,
                "Hora LoRaWAN": None,
                "Data P2P": None,
                "Hora P2P": None,
                "Tem GSM": False,
                "Tem LORAWAN": False,
                "Tem P2P": False
            }
            modelo_counts["N/A"] = modelo_counts.get("N/A", 0) + 1

        adicionar_log(f"ℹ️ Processando {len(resultados)} registros...")

        # Processa resultados
        for idx, resultado in enumerate(resultados):
            try:
                if not isinstance(resultado, dict):
                    continue

                serial = (resultado.get("Serial") or
                         resultado.get("serial") or
                         resultado.get("rastreador_numero_serie"))

                tipo_raw = (resultado.get("Tipo") or
                           resultado.get("tipo") or
                           resultado.get("tipo_evento") or "")
                tipo = str(tipo_raw).lower().strip()

                modelo_hw = (resultado.get("Modelo de HW") or
                            resultado.get("modelo_hw") or
                            resultado.get("rastreador_versao_hardware") or "N/A")

                dados_raw = (resultado.get("Dados", "") or
                            resultado.get("raw_package") or
                            resultado.get("raw") or "")

                data_evt = (resultado.get("DataHora Evento") or
                           resultado.get("data_hora_evento") or
                           resultado.get("data_hora_recebimento") or
                           resultado.get("data_hora_armazenamento"))

                if serial is None:
                    continue

                # Adiciona serial se não estava na lista
                if serial not in serial_data:
                    serial_data[serial] = {
                        "Modelo de HW": "N/A",
                        "dt_gsm": None,
                        "dt_lora": None,
                        "dt_p2p": None,
                        "Data GSM": None,
                        "Hora GSM": None,
                        "Data LoRaWAN": None,
                        "Hora LoRaWAN": None,
                        "Data P2P": None,
                        "Hora P2P": None,
                        "Tem GSM": False,
                        "Tem LORAWAN": False,
                        "Tem P2P": False
                    }
                    serials_list.append(serial)
                    modelo_counts["N/A"] = modelo_counts.get("N/A", 0) + 1

                # Atualiza modelo de HW
                current_model = serial_data[serial]["Modelo de HW"]
                if modelo_hw != "N/A" and current_model == "N/A":
                    serial_data[serial]["Modelo de HW"] = modelo_hw
                    modelo_counts["N/A"] = max(0, modelo_counts.get("N/A", 0) - 1)
                    modelo_counts[modelo_hw] = modelo_counts.get(modelo_hw, 0) + 1

                # Normaliza data usando função centralizada (retorna datetime ou None)
                dt_obj = parse_date_value(data_evt)

                # Classifica por tipo
                if "gsm" in tipo:
                    # guarda o datetime mais recente para o serial (se houver múltiplos)
                    existing = serial_data[serial].get("dt_gsm")
                    if dt_obj and (existing is None or dt_obj > existing):
                        serial_data[serial]["dt_gsm"] = dt_obj
                        serial_data[serial]["Data GSM"] = dt_obj.strftime("%d/%m/%Y")
                        serial_data[serial]["Hora GSM"] = dt_obj.strftime("%H:%M:%S")
                    serial_data[serial]["Tem GSM"] = True
                    dados_gsm[serial] = dados_raw
                    found_serials["gsm"].add(serial)

                elif "lora" in tipo or "lorawan" in tipo:
                    existing = serial_data[serial].get("dt_lora")
                    if dt_obj and (existing is None or dt_obj > existing):
                        serial_data[serial]["dt_lora"] = dt_obj
                        serial_data[serial]["Data LoRaWAN"] = dt_obj.strftime("%d/%m/%Y")
                        serial_data[serial]["Hora LoRaWAN"] = dt_obj.strftime("%H:%M:%S")
                    serial_data[serial]["Tem LORAWAN"] = True
                    dados_lorawan[serial] = dados_raw
                    found_serials["lorawan"].add(serial)

                elif "p2p" in tipo:
                    existing = serial_data[serial].get("dt_p2p")
                    if dt_obj and (existing is None or dt_obj > existing):
                        serial_data[serial]["dt_p2p"] = dt_obj
                        serial_data[serial]["Data P2P"] = dt_obj.strftime("%d/%m/%Y")
                        serial_data[serial]["Hora P2P"] = dt_obj.strftime("%H:%M:%S")
                    serial_data[serial]["Tem P2P"] = True
                    dados_p2p[serial] = dados_raw
                    found_serials["p2p"].add(serial)

            except Exception as e:
                adicionar_log(f"⚠️ Erro processando registro #{idx}: {e}")
                continue

        total_seriais = len(serials_list)
        adicionar_log(f"ℹ️ Total de seriais: {total_seriais}")

        # Calcula valores esperados
        expected = {"gsm": 0, "lorawan": 0, "p2p": 0}
        for s in serials_list:
            modelo = serial_data.get(s, {}).get("Modelo de HW", "N/A")
            regras = REGRAS_HW.get(modelo)
            if regras:
                for req in regras.get("required", []):
                    if req in expected:
                        expected[req] += 1

        # =====================================================================
        # CRIA ABAS DETALHADAS E POR PERÍODO (para depois contabilizar os períodos)
        # =====================================================================

        # Cria abas detalhadas (sempre)
        criar_aba_detalhada_ordenada(wb, "GSM_Detalhado", dados_gsm, 'redis')
        criar_aba_detalhada_ordenada(wb, "LoRaWAN_Detalhado", dados_lorawan, 'redis')
        criar_aba_detalhada_ordenada(wb, "P2P_Detalhado", dados_p2p, 'redis')

        # Obtém configuração de abas do app_state
        app_state = AppState()
        sheet_config = app_state.get("sheet_config", {})

        # Tipos de comunicação: se vier do diálogo, usa; se não, usa todos
        comm_types = sheet_config.get("comm_types")
        if comm_types is None:
            comm_types = ["GSM", "LoRaWAN", "P2P"]
            adicionar_log(f"⚠️ Sem tipos do usuário, usando todos: {comm_types}")
        else:
            adicionar_log(f"📊 Tipos do usuário: {comm_types}")

        # Períodos: se veio como parâmetro (selected_periods), ele é a fonte da verdade
        if selected_periods is not None:
            periods = selected_periods
            adicionar_log(f"📊 Períodos recebidos por parâmetro: {periods}")
        else:
            periods = sheet_config.get("periods")
            if periods is None:
                periods = ["Hoje", "1-7", "8-15", "+16"]
                adicionar_log(f"⚠️ Sem períodos do usuário, usando todos: {periods}")
            else:
                adicionar_log(f"📊 Períodos do usuário (sheet_config): {periods}")

        # CRIA ABAS POR PERÍODO APENAS PARA TIPOS SELECIONADOS
        if "GSM" in comm_types and dados_gsm:
            adicionar_log(f"📊 Criando abas GSM para períodos: {periods}")
            criar_abas_por_periodo(wb, "GSM", dados_gsm, 'redis', periods)
        else:
            if "GSM" not in comm_types:
                adicionar_log("ℹ️ GSM não selecionado pelo usuário")
            else:
                adicionar_log("ℹ️ GSM sem dados")

        if "LoRaWAN" in comm_types and dados_lorawan:
            adicionar_log(f"📊 Criando abas LoRaWAN para períodos: {periods}")
            criar_abas_por_periodo(wb, "LoRaWAN", dados_lorawan, 'redis', periods)
        else:
            if "LoRaWAN" not in comm_types:
                adicionar_log("ℹ️ LoRaWAN não selecionado pelo usuário")
            else:
                adicionar_log("ℹ️ LoRaWAN sem dados")

        if "P2P" in comm_types and dados_p2p:
            adicionar_log(f"📊 Criando abas P2P para períodos: {periods}")
            criar_abas_por_periodo(wb, "P2P", dados_p2p, 'redis', periods)
        else:
            if "P2P" not in comm_types:
                adicionar_log("ℹ️ P2P não selecionado pelo usuário")
            else:
                adicionar_log("ℹ️ P2P sem dados")

        # =====================================================================
        # ABA RESUMO_ULTIMA_POSICAO
        # =====================================================================

        # Funções utilitárias
        def calc_rel(found, expect):
            """Encontrada/Esperada [%]"""
            return (found / expect * 100) if expect else 0.0

        def calc_abs(found, total):
            """Posição Hoje [%] ou similar"""
            return (found / total * 100) if total else 0.0

        # Contadores únicos por tecnologia (encontrado)
        contagem = {
            "gsm": len(found_serials["gsm"]),
            "lorawan": len(found_serials["lorawan"]),
            "p2p": len(found_serials["p2p"])
        }

        # Quantidade sem comunicação (seriais que não aparecem em nenhum conjunto)
        sem_comunic = len([s for s in serials_list
                          if s not in found_serials["gsm"]
                          and s not in found_serials["lorawan"]
                          and s not in found_serials["p2p"]])

        # Total válidos = inseridos - sem comunicação
        total_validos = total_seriais - sem_comunic

        # =====================================================================
        # BLOCO 1: TOTAIS (linhas 1-3)
        # =====================================================================

        ws_main.append(["Total de Seriais Inseridos", total_seriais])
        ws_main.append(["Total de Seriais Válidos", total_validos, f"{(total_validos/total_seriais*100):.1f}%" if total_seriais else "0%"])
        ws_main.append(["Sem comunicação", sem_comunic, f"{(sem_comunic/total_seriais*100):.1f}%" if total_seriais else "0%"])

        # Formata bloco de totais (negrito)
        for row_idx in range(1, 4):
            for col_idx in range(1, 4):
                cell = ws_main.cell(row=row_idx, column=col_idx)
                cell.font = Font(bold=True)

        # Linha em branco
        ws_main.append([])

        # =====================================================================
        # BLOCO 2: TIPO DE COMUNICAÇÃO (linhas 5+)
        # =====================================================================

        # Função para contar quantos seriais há em uma aba específica criada (se existir)
        def count_entries_in_sheet(wb_obj, sheet_name):
            if sheet_name not in wb_obj.sheetnames:
                return 0
            ws = wb_obj[sheet_name]
            count = 0
            # procura na primeira coluna, a partir da linha 2
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
                if r and r[0] not in (None, ""):
                    count += 1
            return count

        # Mapeia os nomes das abas de período que provavelmente foram criadas por criar_abas_por_periodo
        def build_period_sheet_name(prefix, period_label):
            label = str(period_label).strip()
            return f"{prefix}_{label}"

        # Obtem contagens por período para um dado prefixo (GSM, LoRaWAN, P2P)
        def get_period_counts_for_prefix(prefix):
            counts = []
            for p in periods:
                sheet_name = build_period_sheet_name(prefix, p)
                cnt = count_entries_in_sheet(wb, sheet_name)
                counts.append(cnt)
            return counts

        # Prepara contagens por período
        gsm_counts_by_period = get_period_counts_for_prefix("GSM")
        lorawan_counts_by_period = get_period_counts_for_prefix("LoRaWAN")
        p2p_counts_by_period = get_period_counts_for_prefix("P2P")

        # Cabeçalho do bloco de comunicação
        header_comm = ["Tipo de Comunicação", "Quantidade Encontrada", "Quantidade Esperada", "Encontrada/Esperada [%]"]

        # Adiciona cabeçalhos de períodos
        for p in periods:
            header_comm.append(f"Posição {p}")
            header_comm.append(f"Posição {p} [%]")

        ws_main.append(header_comm)
        header_comm_row = ws_main.max_row

        # Formata cabeçalho de comunicação (negrito + centralizado)
        for col_idx in range(1, len(header_comm) + 1):
            cell = ws_main.cell(row=header_comm_row, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Dados de comunicação
        # GSM
        gsm_encontrada_esperada = calc_rel(contagem["gsm"], expected["gsm"])
        gsm_row = ["GSM", contagem["gsm"], expected["gsm"], f"{gsm_encontrada_esperada:.1f}%"]
        for idx, p in enumerate(periods):
            gsm_row.append(gsm_counts_by_period[idx])
            gsm_row.append(f"{calc_abs(gsm_counts_by_period[idx], contagem['gsm']):.2f}%" if contagem["gsm"] > 0 else "0%")
        ws_main.append(gsm_row)

        # LoRaWAN
        lorawan_encontrada_esperada = calc_rel(contagem["lorawan"], expected["lorawan"])
        lorawan_row = ["LoRaWAN", contagem["lorawan"], expected["lorawan"], f"{lorawan_encontrada_esperada:.1f}%"]
        for idx, p in enumerate(periods):
            lorawan_row.append(lorawan_counts_by_period[idx])
            lorawan_row.append(f"{calc_abs(lorawan_counts_by_period[idx], contagem['lorawan']):.2f}%" if contagem["lorawan"] > 0 else "0%")
        ws_main.append(lorawan_row)

        # P2P
        p2p_encontrada_esperada = calc_rel(contagem["p2p"], expected["p2p"])
        p2p_row = ["P2P", contagem["p2p"], expected["p2p"], f"{p2p_encontrada_esperada:.1f}%"]
        for idx, p in enumerate(periods):
            p2p_row.append(p2p_counts_by_period[idx])
            p2p_row.append(f"{calc_abs(p2p_counts_by_period[idx], contagem['p2p']):.2f}%" if contagem["p2p"] > 0 else "0%")
        ws_main.append(p2p_row)

        # Linha em branco
        ws_main.append([])

        # =====================================================================
        # TABELA PRINCIPAL DE SERIAIS
        # =====================================================================

        # Tabela principal: cabeçalhos com Data/Hora separados
        headers = [
            "Serial", "Modelo de HW", "Posição GSM",
            "Data GSM", "Hora GSM",
            "Posição LoRaWAN", "Data LoRaWAN", "Hora LoRaWAN",
            "Posição P2P", "Data P2P", "Hora P2P"
        ]
        ws_main.append(headers)
        main_header_row = ws_main.max_row

        # Ordena serials: prioridade dt_gsm (desc), se não há, dt_p2p, se não, dt_lora
        def sort_key_for_serial(s):
            info = serial_data.get(s, {})
            # prefer GSM, then P2P, then LoRaWAN
            dt = info.get("dt_gsm") or info.get("dt_p2p") or info.get("dt_lora")
            # retornamos uma tupla (has_dt, dt) para ordenar; sem dt -> menor
            if dt:
                return (1, dt)
            else:
                return (0, datetime.min)

        serials_sorted = sorted(serials_list, key=sort_key_for_serial, reverse=True)

        # Helper to get position status baseado em regras e flags
        def get_pos_status(has_flag, modelo, proto):
            if has_flag:
                return "SIM"
            if not REGRAS_HW or not modelo:
                return "NÃO"
            regras = REGRAS_HW.get(modelo)
            if not regras:
                return "NÃO"
            if proto in regras.get("required", []) or proto in regras.get("optional", []):
                return "NÃO"
            return "NÃO POSSUI"

        for s in serials_sorted:
            info = serial_data.get(s, {})
            modelo = info.get("Modelo de HW", "N/A")

            row = [
                s,
                modelo,
                get_pos_status(info.get("Tem GSM"), modelo, "gsm"),
                info.get("Data GSM") or "N/A",
                info.get("Hora GSM") or "N/A",
                get_pos_status(info.get("Tem LORAWAN"), modelo, "lorawan"),
                info.get("Data LoRaWAN") or "N/A",
                info.get("Hora LoRaWAN") or "N/A",
                get_pos_status(info.get("Tem P2P"), modelo, "p2p"),
                info.get("Data P2P") or "N/A",
                info.get("Hora P2P") or "N/A"
            ]
            ws_main.append(row)

        # Formata aba principal usando funções centralizadas
        # Formata cabeçalho usando função do report_utils
        formatar_cabecalho(ws_main, main_header_row)

        # Aplica zebra apenas na tabela principal (a partir de main_header_row + 1)
        #aplicar_estilo_zebra(ws_main, main_header_row + 1)

        # Congela a linha do cabeçalho
        ws_main.freeze_panes = f"A{main_header_row + 1}"

        # Aplica filtro automático
        if ws_main.max_column >= 1:
            last_col = get_column_letter(ws_main.max_column)
            ws_main.auto_filter.ref = f"A{main_header_row}:{last_col}{ws_main.max_row}"

        # Ajusta largura das colunas
        ajustar_largura_colunas(ws_main)

        # =====================================================================
        # ABA RESUMO_MODELO_HW (NOVA - SEGUNDA POSIÇÃO)
        # =====================================================================

        ws_modelo = wb.create_sheet("Resumo_Modelo_HW")

        # Função para contar quantos seriais de um modelo específico estão em uma aba de período
        def count_model_in_period_sheet(wb_obj, sheet_name, modelo_target):
            """
            Conta quantos seriais do modelo especificado aparecem na aba de período.
            """
            if sheet_name not in wb_obj.sheetnames:
                return 0

            ws = wb_obj[sheet_name]
            count = 0

            # Procura na primeira coluna (Serial), a partir da linha 2
            for row_idx in range(2, ws.max_row + 1):
                serial_cell = ws.cell(row=row_idx, column=1).value
                if serial_cell and serial_cell in serial_data:
                    modelo_serial = serial_data[serial_cell].get("Modelo de HW", "N/A")
                    if modelo_serial == modelo_target:
                        count += 1

            return count

        # Cabeçalho do bloco de modelo
        header_modelo = ["Modelo de HW", "Quantidade encontrada", "Quantidade Encontrada [%]"]

        # Adiciona cabeçalhos de períodos baseados nos tipos de comunicação selecionados
        for comm_type in comm_types:
            for p in periods:
                header_modelo.append(f"Posição {comm_type} {p}")
                header_modelo.append(f"Posição {comm_type} {p} [%]")

        ws_modelo.append(header_modelo)
        header_modelo_row = ws_modelo.max_row

        # Dados de modelo
        for modelo in sorted(modelo_counts.keys(), key=lambda s: str(s).lower()):
            modelo_total = modelo_counts[modelo]
            modelo_pct = (modelo_total / total_seriais * 100) if total_seriais else 0

            modelo_row = [modelo, modelo_total, f"{modelo_pct:.2f}%"]

            # Para cada tipo de comunicação selecionado
            for comm_type in comm_types:
                # Mapeia tipo para prefixo de aba
                if comm_type == "GSM":
                    prefix = "GSM"
                elif comm_type == "LoRaWAN":
                    prefix = "LoRaWAN"
                elif comm_type == "P2P":
                    prefix = "P2P"
                else:
                    continue

                # Para cada período, conta quantos seriais deste modelo estão naquele período
                for p in periods:
                    sheet_name = build_period_sheet_name(prefix, p)
                    periodo_count = count_model_in_period_sheet(wb, sheet_name, modelo)
                    periodo_pct = (periodo_count / modelo_total * 100) if modelo_total > 0 else 0

                    modelo_row.append(periodo_count)
                    modelo_row.append(f"{periodo_pct:.2f}%")

            ws_modelo.append(modelo_row)

        # Formata aba de modelo usando funções centralizadas
        # Formata cabeçalho usando função do report_utils
        formatar_cabecalho(ws_modelo, header_modelo_row)

        # Aplica zebra a partir da linha após o cabeçalho
        #aplicar_estilo_zebra(ws_modelo, header_modelo_row + 1)

        # Congela APENAS a primeira coluna (coluna A) para rolagem horizontal
        ws_modelo.freeze_panes = "B1"

        # Aplica filtro automático
        if ws_modelo.max_column >= 1:
            last_col = get_column_letter(ws_modelo.max_column)
            ws_modelo.auto_filter.ref = f"A{header_modelo_row}:{last_col}{ws_modelo.max_row}"

        # Ajusta largura das colunas
        ajustar_largura_colunas(ws_modelo)

        # Reposiciona a aba Resumo_Modelo_HW para segunda posição
        try:
            if ws_modelo in wb._sheets:
                wb._sheets.remove(ws_modelo)
                wb._sheets.insert(1, ws_modelo)
        except Exception:
            pass


        # =====================================================================
        # ABA EQUIP_SEM_POSICAO (TERCEIRA POSIÇÃO)
        # =====================================================================

        ws_sem = wb.create_sheet("Equip_sem_posicao")
        ws_sem.append(["Serial"])
        sem_list = []

        for s in serials_list:
            info = serial_data.get(s, {})
            if not info.get("Tem GSM") and not info.get("Tem LORAWAN") and not info.get("Tem P2P"):
                sem_list.append(s)

        # Ordena sem_list do menor para o maior (tenta int -> fallback str)
        def sort_sem_key(x):
            try:
                return int(x)
            except Exception:
                return str(x)

        sem_list_sorted = sorted(sem_list, key=sort_sem_key)

        for s in sem_list_sorted:
            ws_sem.append([s])

        # Formata usando função centralizada
        formatar_planilha_completa(ws_sem)

        # Reposiciona a aba sem_posicao para terceira posição
        try:
            if ws_sem in wb._sheets:
                wb._sheets.remove(ws_sem)
                wb._sheets.insert(2, ws_sem)
        except Exception:
            pass

        # =====================================================================
        # Salva arquivo
        # =====================================================================

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        adicionar_log(f"✅ Relatório Redis salvo em: {os.path.abspath(output_path)}")

        return output_path

    except Exception as e:
        adicionar_log(f"❌ Erro em gerar_relatorio_redis: {e}")
        adicionar_log(traceback.format_exc())
        return None

# =============================================================================
# GERAÇÃO ESPECÍFICA PARA API
# =============================================================================

def gerar_relatorio_api(serials, resultados, output_path):
    """
    Gera relatório simplificado para dados da API Mogno.
    Mantém formatação original da API.

    Args:
        serials: Lista de seriais solicitados
        resultados: Lista de registros com dados
        output_path: Caminho completo para salvar o arquivo
    """
    try:
        adicionar_log("📡 [API] Iniciando geração do relatório...")

        if not resultados:
            adicionar_log("⚠️ Nenhum dado disponível para gerar relatório (API).")
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Ultimas_Posicoes_API"

        # Cabeçalhos e dados
        if isinstance(resultados, list) and resultados:
            if isinstance(resultados[0], dict):
                headers = list(resultados[0].keys())
            else:
                headers = ["Serial", "Valor"]
        else:
            headers = ["Dados"]

        ws.append(headers)

        # Normaliza valores antes de escrever
        for item in resultados:
            if isinstance(item, dict):
                row = []
                for value in item.values():
                    # Converte objetos não suportados para string
                    if isinstance(value, (dict, list)):
                        value = str(value)
                    elif not isinstance(value, (str, int, float, bool, type(None))):
                        # Converte enums, protobuf, etc. para string
                        value = str(value)

                    # Trunca strings muito longas
                    if isinstance(value, str) and len(value) > 32000:
                        value = value[:32000] + "… [TRUNCADO]"

                    row.append(value)

                ws.append(row)
            else:
                ws.append([str(item)])

        # Formata usando função centralizada
        formatar_planilha_completa(ws)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        adicionar_log(f"✅ Relatório API salvo em: {os.path.relpath(output_path)}")

        return output_path

    except Exception as e:
        adicionar_log(f"❌ Erro em gerar_relatorio_api: {e}")
        adicionar_log(traceback.format_exc())
        return None

# =============================================================================
# FUNÇÃO PRINCIPAL UNIFICADA (INTERFACE PÚBLICA)
# =============================================================================

def gerar_relatorio(serials, resultados, output_path, selected_periods=None, origem='redis'):
    """
    Interface unificada para geração de relatórios de Últimas Posições.
    A origem é determinada pelo tipo de requisição (não há detecção automática).

    Args:
        serials: Lista de seriais solicitados
        resultados: Lista de registros com dados
        output_path: Caminho completo para salvar o arquivo
        selected_periods: Lista de períodos a incluir (ex: ["Hoje", "1-7"])
        origem: 'redis' ou 'api' (obrigatório)

    Returns:
        str: Caminho do arquivo gerado ou None em caso de erro
    """
    try:
        if not resultados and not serials:
            adicionar_log("⚠️ Nenhum dado disponível para gerar relatório.")
            return None

        # Usa origem explícita
        adicionar_log(f"🔍 Origem: {origem.upper()}")

        # Delega para função específica
        if origem == 'redis':
            return gerar_relatorio_redis(serials, resultados, output_path, selected_periods)
        else:
            return gerar_relatorio_api(serials, resultados, output_path)

    except Exception as e:
        adicionar_log(f"❌ Erro crítico em gerar_relatorio: {e}")
        adicionar_log(traceback.format_exc())
        return None
