"""
Gerador de relatório Excel para análise de eventos de rastreadores.
Versão final estabilizada e otimizada para grandes volumes (>500k eventos).
"""
import re
import traceback
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config.settings import EVENT_NAMES
from utils.logger import adicionar_log
from utils.helpers import calcular_periodo_dias
from reports.reports_utils import (
    parse_date_value,
    formatar_cabecalho_customizado,
    ajustar_largura_colunas_otimizado,
    formatar_planilha_modo_rapido,
    formatar_planilha_completa
)

# =============================================================================
# PARSER PROTO
# =============================================================================

def parse_proto_eventos(proto_str):
    """
    Parser de proto textual:
    - Respeita hierarquia real com {}
    - Flatten com "_" (rastreador_bateria_interna_tensao)
    - Bloqueia horário solto virar chave (16:00:00)
    - Preserva data + hora corretamente
    """

    if not proto_str:
        return {}



    proto_str = re.sub(r'<[^>]*>', ' ', proto_str)
    proto_str = re.sub(r'\s+', ' ', proto_str).strip()

    token_pattern = re.compile(
        r'''
        [A-Za-z_][\w\-]*\s*\{ |
        \} |
        [A-Za-z_][\w\-]*:\s*"[^"]*" |
        [A-Za-z_][\w\-]*:\s*\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2} |
        [A-Za-z_][\w\-]*:\s*[^{}\s]+
        ''',
        re.VERBOSE
    )

    tokens = token_pattern.findall(proto_str)

    resultado = {}
    stack = []
    block_count = defaultdict(int)

    for token in tokens:
        token = token.strip()

        if token.endswith("{"):
            bloco = token[:-1].strip()
            level_key = (tuple(stack), bloco)
            block_count[level_key] += 1
            idx = block_count[level_key]
            stack.append(f"{bloco}_{idx}" if idx > 1 else bloco)
            continue

        if token == "}":
            if stack:
                stack.pop()
            continue

        if ":" in token:
            chave, valor = token.split(":", 1)
            chave = chave.strip()
            valor = valor.strip().strip('"')

            if valor.lower() in ("true", "false"):
                valor = valor.lower() == "true"

            elif re.match(r"\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}", valor):
                pass

            else:
                try:
                    valor = int(valor)
                except:
                    try:
                        valor = float(valor)
                    except:
                        pass

            chave_final = "_".join(stack + [chave]) if stack else chave
            resultado[chave_final] = valor

    return resultado

# =============================================================================
# PROCESSAMENTO DE EVENTOS
# =============================================================================

def extrair_tipo_evento(proto_str):
    """Extrai tipo de evento do campo 'proto' (ex: 'IGNICAO_ON')."""
    try:
        match = re.search(r'tipo_evento:\s*(\w+)', proto_str)
        return match.group(1) if match else "DESCONHECIDO"
    except:
        return "DESCONHECIDO"


def _processar_eventos(eventos_data):
    """
    Processa eventos brutos e extrai informações estruturadas.

    Inclui campos da API (horario_api) para garantir datas no resumo
    """
    eventos_parseados = []
    tipos_eventos = set()
    seriais_com_eventos = set()

    for evento in eventos_data:
        try:
            proto = evento.get("proto") or ""
            serial_api = evento.get("serial")

            # Parse do proto
            dados_proto = parse_proto_eventos(proto)

            if not dados_proto:
                continue

            # Serial: prioriza API, fallback para proto
            serial = serial_api if serial_api and serial_api != "N/A" else dados_proto.get("rastreador_numero_serie")

            if not serial:
                continue

            # Tipo de evento
            tipo_evento = dados_proto.get("tipo_evento") or extrair_tipo_evento(proto)
            tipos_eventos.add(tipo_evento)
            seriais_com_eventos.add(serial)

            # Monta evento final com CAMPOS DA API (crítico para o resumo funcionar)
            evento_final = {
                "_serial": serial,
                "_tipo_evento": tipo_evento
            }

            # CAMPOS DA API (garantem que horario_api sempre exista)
            horario_api = evento.get("horario") or evento.get("data_hora_evento") or ""
            evento_final["horario_api"] = horario_api
            
            # Campos do proto (cada chave:valor em coluna separada)
            evento_final.update(dados_proto)

            eventos_parseados.append(evento_final)

        except Exception as e:
            adicionar_log(f"⚠️ Evento ignorado por erro: {e}")
            continue

    return eventos_parseados, sorted(tipos_eventos), seriais_com_eventos

# =============================================================================
# DADOS DE RESUMO DE EVENTOS
# =============================================================================

def _gerar_resumo_eventos(eventos_parseados):
    """
    Gera dados consolidados para aba de resumo.
    - Prioriza horario_api (sempre existe agora)
    - Fallback para campos do proto com prefixos

    Returns:
        dict: {serial: {tipo_evento: {'ultima_data': datetime, 'quantidade': int}}}
    """
    resumo = defaultdict(lambda: defaultdict(lambda: {
        "ultima_data": None,
        "quantidade": 0
    }))

    for ev in eventos_parseados:
        serial = ev.get("_serial")
        tipo = ev.get("_tipo_evento")

        if not serial or not tipo:
            continue

        resumo[serial][tipo]["quantidade"] += 1

        # PRIORIDADE 1: horario_api (vem da API, sempre confiável)
        horario_api = ev.get("horario_api")
        if horario_api and horario_api != "N/A":
            try:
                dt = datetime.strptime(horario_api, "%d/%m/%Y %H:%M:%S")
                atual = resumo[serial][tipo]["ultima_data"]
                if atual is None or dt > atual:
                    resumo[serial][tipo]["ultima_data"] = dt
                continue  # ✅ Encontrou, pula para próximo evento
            except:
                pass

        # FALLBACK: busca em campos do proto (com ou sem prefixos)
        for campo in ("data_hora_evento", "data_hora_recebimento", "data_hora_armazenamento"):
            for k, v in ev.items():
                if campo in k:  # Aceita gps_0_data_hora_evento, etc.
                    dt = parse_date_value(v)
                    if dt:
                        atual = resumo[serial][tipo]["ultima_data"]
                        if atual is None or dt > atual:
                            resumo[serial][tipo]["ultima_data"] = dt
                    break
            else:
                continue
            break  # Encontrou data válida, para de procurar

    return resumo

# =============================================================================
# ABA RESUMO
# =============================================================================

def _criar_aba_resumo(wb, resumo_data, eventos_parseados, serials, start_datetime, end_datetime, filtros_list):
    """
    Cria aba "Resumo_Eventos" com layout executivo + seção de modelos + tabela detalhada.

    Estrutura:
    - Linhas 1-3: Estatísticas gerais
    - Linha 4: (em branco)
    - Linha 5: Período de avaliação
    - Linha 6: (em branco)
    - Linha 7: Cabeçalho da seção de modelos
    - Linhas 8+: Dados por modelo
    - Linha N: (em branco)
    - Linha N+1: Cabeçalho da tabela principal (com filtro e congelamento)
    - Linha N+2+: Dados detalhados por serial

    Args:
        wb: Workbook do openpyxl
        resumo_data: dict {serial: {tipo_evento: {'ultima_data': datetime, 'quantidade': int}}}
        eventos_parseados: lista de eventos com dados do proto
        serials: lista de seriais requisitados
        start_datetime: data/hora início (string)
        end_datetime: data/hora fim (string)
        filtros_list: lista de tipos de eventos (filtros)
    """
    try:
        ws = wb.create_sheet("Resumo_Eventos", 0)

        # ========== CÁLCULOS PRELIMINARES ==========
        total_serials = len(serials)

        # Seriais válidos: aqueles que têm pelo menos 1 evento
        valid_serials = [s for s in serials if s in resumo_data and resumo_data[s]]
        num_valid = len(valid_serials)

        # Seriais sem eventos
        num_sem_eventos = total_serials - num_valid

        # Percentuais
        pct_valid = (num_valid / total_serials * 100) if total_serials > 0 else 0
        pct_sem_eventos = (num_sem_eventos / total_serials * 100) if total_serials > 0 else 0

        # Tipos de eventos realmente encontrados
        tipos_eventos_resumo = sorted({
            tipo
            for serial_data in resumo_data.values()
            for tipo in serial_data.keys()
        })

        if not tipos_eventos_resumo:
            ws.append(["Nenhum evento encontrado"])
            return

        # ========== EXTRAI MODELOS DE DISPOSITIVOS ==========
        # Cria mapa: serial → modelo (rastreador_versao_hardware)
        serial_to_modelo = {}
        for evento in eventos_parseados:
            serial = evento.get("_serial")
            modelo = evento.get("rastreador_versao_hardware", "DESCONHECIDO")
            if serial and serial not in serial_to_modelo:
                serial_to_modelo[serial] = modelo

        # Agrupa seriais por modelo e conta eventos por modelo
        modelos_dict = {}  # {modelo: {"seriais": set(), "eventos": {tipo: set()}}}
        for serial in valid_serials:
            modelo = serial_to_modelo.get(serial, "DESCONHECIDO")

            if modelo not in modelos_dict:
                modelos_dict[modelo] = {
                    "seriais": set(),
                    "eventos": {tipo: set() for tipo in tipos_eventos_resumo}
                }

            modelos_dict[modelo]["seriais"].add(serial)

            # Para cada tipo de evento que este serial teve
            for tipo in resumo_data.get(serial, {}).keys():
                if tipo in modelos_dict[modelo]["eventos"]:
                    modelos_dict[modelo]["eventos"][tipo].add(serial)

        # ========== LINHA 1: Total de seriais inseridos ==========
        ws.append(["Total de seriais Inseridos", total_serials])

        # ========== LINHA 2: Total de seriais válidos ==========
        ws.append(["Total de seriais Válidos", num_valid, f"{pct_valid:.2f}%"])

        # ========== LINHA 3: Total de seriais sem eventos ==========
        ws.append(["Total de seriais sem eventos", num_sem_eventos, f"{pct_sem_eventos:.2f}%"])

        # ========== LINHA 4: Em branco ==========
        ws.append([])

        # ========== LINHA 5: Período de Avaliação ==========
        dias = calcular_periodo_dias(start_datetime, end_datetime)
        ws.append(["Período de Avaliação", start_datetime, end_datetime, f"{dias} dias"])

        # ========== LINHA 6: Em branco ==========
        ws.append([])

        # ========== LINHA 7: CABEÇALHO DA SEÇÃO DE MODELOS ==========
        headers_modelos = ["Modelos de Dispositivos", "Quantidade", "Quantidade [%]"] + tipos_eventos_resumo
        ws.append(headers_modelos)

        # ========== LINHAS 8+: DADOS POR MODELO ==========
        for modelo in sorted(modelos_dict.keys()):
            info = modelos_dict[modelo]
            qtd_modelo = len(info["seriais"])
            pct_modelo = (qtd_modelo / num_valid * 100) if num_valid > 0 else 0

            row_modelo = [modelo, qtd_modelo, f"{pct_modelo:.2f}%"]

            # Para cada tipo de evento, conta quantos seriais deste modelo tiveram aquele evento
            for tipo in tipos_eventos_resumo:
                qtd_tipo_modelo = len(info["eventos"][tipo])
                row_modelo.append(qtd_tipo_modelo)

            ws.append(row_modelo)

        # ========== LINHA VAZIA ==========
        ws.append([])

        # ========== CALCULA LINHA DO CABEÇALHO DA TABELA PRINCIPAL ==========
        linha_cabecalho = ws.max_row + 1

        # ========== CABEÇALHO DA TABELA PRINCIPAL ==========
        headers_tabela = ["Serial", "Modelo de Dispositivo"]
        for tipo in tipos_eventos_resumo:
            headers_tabela.extend([tipo, f"Quantidade_{tipo}"])
        ws.append(headers_tabela)

        linha_cabecalho = ws.max_row
        
        # ========== LINHAS SEGUINTES: DADOS DA TABELA ==========
        for serial in sorted(valid_serials):
            dados = resumo_data.get(serial, {})
            modelo = serial_to_modelo.get(serial, "DESCONHECIDO")

            if not dados:
                continue

            row = [serial, modelo]
            for tipo in tipos_eventos_resumo:
                info = dados.get(tipo)
                if info:
                    dt = info.get("ultima_data")
                    row.append(
                        dt.strftime("%d/%m/%Y %H:%M:%S") if dt else "N/A"
                    )
                    row.append(info.get("quantidade", 0))
                else:
                    row.extend(["N/A", 0])

            ws.append(row)

        # ========== FORMATAÇÃO DO CABEÇALHO DA TABELA PRINCIPAL ==========
        formatar_cabecalho_customizado(ws, linha_cabecalho, "305496")

        # Congela painéis: linhas acima do cabeçalho ficam fixas
        ws.freeze_panes = f"A{linha_cabecalho + 1}"

        # Filtro: aplica na linha do cabeçalho da tabela
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{linha_cabecalho}:{last_col}{ws.max_row}"

        # Ajusta largura das colunas
        ajustar_largura_colunas_otimizado(ws)

    except Exception as e:
        adicionar_log(f"❌ Erro ao criar aba de resumo: {e}")
        adicionar_log(traceback.format_exc())



# =============================================================================
# ABA SEM EVENTOS
# =============================================================================

def criar_aba_seriais_sem_evento(wb, seriais_sem_evento, filtros_list):
    """
    Cria aba "Seriais_sem_evento" com seriais que não tiveram eventos encontrados.

    Args:
        wb: Workbook do openpyxl
        seriais_sem_evento: list de seriais (strings)
        filtros_list: list de nomes de eventos filtrados

    Returns:
        Worksheet criado ou None
    """
    if not seriais_sem_evento:
        adicionar_log("ℹ️ Nenhum serial sem evento para criar aba")
        return None

    try:
        ws = wb.create_sheet("Seriais_sem_evento", 1)  # Posição 1 (segunda aba)

        # Cabeçalho: Seriais + cada filtro
        headers = ["Seriais"] + ["Status"]
        ws.append(headers)

        # Dados: cada serial + "NÃO ENCONTRADO" para cada filtro
        for serial in seriais_sem_evento:
            row = [serial] + ["NENHUM TIPO DE EVENTO ENCONTRADO NO PERÍODO SELECIONADO"]
            ws.append(row)

        # Formata (sem zebra, com ajuste de colunas)
        formatar_planilha_completa(ws, header_row=1, aplicar_zebra=False, ajustar_colunas=True)

        #adicionar_log(f"✅ Aba 'Seriais_sem_evento' criada com {len(seriais_sem_evento)} seriais")
        return ws

    except Exception as e:
        adicionar_log(f"❌ Erro ao criar aba Seriais_sem_evento: {e}")
        adicionar_log(traceback.format_exc())
        return None

# =============================================================================
# ABAS DETALHADAS
# =============================================================================

def _criar_abas_detalhadas(wb, eventos_parseados, tipos_eventos):
    eventos_por_tipo = defaultdict(list)

    for ev in eventos_parseados:
        eventos_por_tipo[ev["_tipo_evento"]].append(ev)

    for tipo in tipos_eventos:
        eventos = eventos_por_tipo.get(tipo)
        if not eventos:
            continue

        ws = wb.create_sheet(tipo[:31])

        headers = []
        headers_set = set()
        for ev in eventos:
            for k in ev.keys():
                if not k.startswith("_") and k not in headers_set:
                    headers_set.add(k)
                    headers.append(k)

        ws.append(headers)

        def extrair_dt(ev):
            for campo in ("data_hora_evento", "data_hora_recebimento", "data_hora_armazenamento"):
                dt = parse_date_value(ev.get(campo))
                if dt:
                    return dt
            return datetime.min

        for ev in sorted(eventos, key=extrair_dt, reverse=True):
            ws.append([ev.get(h, "") for h in headers])

# =============================================================================
# FUNÇÃO PRINCIPAL
# =============================================================================

def gerar_relatorio(
    serials,
    eventos_data,
    output_path,
    start_datetime=None,
    end_datetime=None,
    filtros_str="",
    modo_rapido=True
):
    """
    Gera relatório Excel completo de eventos.

    Args:
        serials: lista de seriais requisitados
        eventos_data: lista de eventos brutos da API
        output_path: caminho do arquivo Excel de saída
        start_datetime: data/hora início (string)
        end_datetime: data/hora fim (string)
        filtros_str: IDs dos filtros separados por vírgula
        modo_rapido: se True, aplica formatação rápida (sem zebra)

    Returns:
        str: caminho do arquivo gerado ou None em caso de erro
    """
    try:
        adicionar_log(f"📊 Gerando relatório: {output_path}")

        # ========== 1. PROCESSA EVENTOS ==========
        eventos_parseados, tipos_eventos, seriais_com_eventos = _processar_eventos(eventos_data)

        if not eventos_parseados:
            adicionar_log("⚠️ Nenhum evento válido processado")
            return None

        # ========== 2. CONVERTE FILTROS ==========
        filtros_list = []
        if filtros_str:
            for fid in filtros_str.split(","):
                try:
                    filtros_list.append(EVENT_NAMES.get(int(fid.strip()), f"Evento_{fid}"))
                except:
                    pass

        if not filtros_list:
            filtros_list = list(tipos_eventos)

        # ========== 3. GERA RESUMO ==========
        resumo_data = _gerar_resumo_eventos(eventos_parseados)

        # ========== 4. CRIA WORKBOOK ==========
        wb = Workbook()
        wb.remove(wb.active)

        # ========== 5. CRIA ABA RESUMO (com formatação customizada interna) ==========
        # adiciona 'serials' como parâmetro
        _criar_aba_resumo(wb, resumo_data, eventos_parseados, serials, start_datetime, end_datetime, filtros_list)

        # ========== 6. CRIA ABA SERIAIS SEM EVENTO ==========
        seriais_sem_evento = sorted(set(serials) - seriais_com_eventos)
        if seriais_sem_evento:
            criar_aba_seriais_sem_evento(wb, seriais_sem_evento, filtros_list)

        # ========== 7. CRIA ABAS DETALHADAS ==========
        _criar_abas_detalhadas(wb, eventos_parseados, tipos_eventos)

        # ========== 8. FORMATAÇÃO FINAL (EXCETO RESUMO_EVENTOS) ==========
        # pula a aba "Resumo_Eventos" para não interferir na formatação customizada
        for ws in wb.worksheets:
            if ws.title == "Resumo_Eventos":
                continue  # Pula formatação automática nesta aba

            if modo_rapido:
                formatar_planilha_modo_rapido(ws)
                ajustar_largura_colunas_otimizado(ws)
            else:
                formatar_planilha_completa(ws)

        # ========== 9. SALVA ARQUIVO ==========
        wb.save(output_path)

        adicionar_log("✅ Relatório gerado com sucesso")
        adicionar_log(f"   📊 Total de eventos: {len(eventos_parseados)}")
        adicionar_log(f"   📋 Tipos de eventos: {len(tipos_eventos)}")
        adicionar_log(f"   ✅ Seriais com eventos: {len(seriais_com_eventos)}")
        adicionar_log(f"   ❌ Seriais sem eventos: {len(seriais_sem_evento)}")

        return output_path

    except Exception as e:
        adicionar_log(f"❌ Erro ao gerar relatório: {e}")
        adicionar_log(traceback.format_exc())
        return None

