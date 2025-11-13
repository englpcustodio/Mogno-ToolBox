# reports/report_last_position_API.py
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from utils.logger import adicionar_log

def gerar_relatorio(serials, resultados, output_dir, separado=True):
    """Gera relatório de Últimas Posições (via API Mogno)"""
    try:
        if not resultados:
            adicionar_log("⚠️ Nenhum dado disponível para gerar relatório (API).")
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Ultimas_Posicoes_API"

        # Cabeçalhos e dados
        headers = list(resultados[0].keys()) if isinstance(resultados, list) else ["Serial", "Valor"]
        ws.append(headers)
        for item in resultados:
            if isinstance(item, dict):
                ws.append(list(item.values()))
            else:
                ws.append([item])

        _formatar_planilha(ws)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_path = os.path.join(output_dir, f"relatorio_ultimas_posicoes_API_{timestamp}.xlsx")
        wb.save(output_path)
        adicionar_log(f"✅ Relatório de Últimas Posições (API) salvo em: {output_path}")
        return output_path

    except Exception as e:
        adicionar_log(f"❌ Erro ao gerar relatório API: {e}")
        return None


def _formatar_planilha(ws):
    """Aplica formatação padrão (premium)"""
    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Cabeçalho
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Corpo
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for c in row:
                c.fill = gray_fill
        for c in row:
            c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Largura automática e congelar
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3
    ws.freeze_panes = "A2"
